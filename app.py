from flask import Flask, render_template, request, redirect, session, url_for, flash, jsonify, send_file
from functools import wraps
from datetime import datetime, timedelta
from io import StringIO, BytesIO
import csv, os, json, zipfile
from utils.json_db import load_json, save_json
import openai
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = "change-this-in-production"

DATA_DIR = "data"
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

def data_path(name): return f"{DATA_DIR}/{name}"
def get_users(): return load_json(data_path("users.json"), [])
def get_teams(): return load_json(data_path("teams.json"), [])
def get_employees(): return load_json(data_path("employees.json"), [])
def get_projects(): return load_json(data_path("projects.json"), [])
def get_billing_rates(): return load_json(data_path("billing_rates.json"), [])
def get_cost_rates(): return load_json(data_path("cost_rates.json"), [])
def get_staffing_entries(): return load_json(data_path("staffing_entries.json"), [])
def get_billing_entries(): return load_json(data_path("billing_entries.json"), [])
def get_reports(): return load_json(data_path("reports.json"), [])
def get_insync_employee_orders(): return load_json(data_path("insync_employee_orders.json"), {})
def get_dropdown_options(): return load_json(data_path("dropdown_options.json"), {})

def save_staffing_entries(rows): save_json(data_path("staffing_entries.json"), rows)
def save_billing_entries(rows): save_json(data_path("billing_entries.json"), rows)
def save_projects(rows): save_json(data_path("projects.json"), rows)
def save_users(rows): save_json(data_path("users.json"), rows)
def save_teams(rows): save_json(data_path("teams.json"), rows)
def save_employees(rows): save_json(data_path("employees.json"), rows)
def save_billing_rates(rows): save_json(data_path("billing_rates.json"), rows)
def save_cost_rates(rows): save_json(data_path("cost_rates.json"), rows)
def save_reports(rows): save_json(data_path("reports.json"), rows)
def save_insync_employee_orders(rows): save_json(data_path("insync_employee_orders.json"), rows)
def save_dropdown_options(rows): save_json(data_path("dropdown_options.json"), rows)

FULL_PROJECT_FIELDS = [
    "project_name","project_type","type_for_util","billing_case_code","client_case_code","work_description","product","requestor",
    "nps_contact","case_status","nps_status","case_manager_for_nps","bcn_case_execution_location",
    "billed_to_end_client","case_type","office","case_manager_principal","client_name","case_partner","industry","capability",
    "master_project_name","date_of_request","case_delivery_primary_location","outside_bcn_location","case_poc","end_client_poc","team_id","region"
]

def resolve_staffing_team_id(staffing_row, project_map=None, employee_map=None):
    project_map = project_map or {}
    employee_map = employee_map or {}
    project_id = staffing_row.get("project_id", "")
    if project_id and project_id in project_map:
        project_team_id = project_map[project_id].get("team_id", "")
        if project_team_id:
            return project_team_id
    if staffing_row.get("team_id", ""):
        return staffing_row.get("team_id", "")
    employee = employee_map.get(staffing_row.get("employee_id", ""), {})
    return employee.get("team_id", "")

def next_id(prefix, rows):
    max_num = 0
    for row in rows:
        value = row.get("id", "")
        if value.startswith(prefix):
            try: max_num = max(max_num, int(value.replace(prefix, "")))
            except ValueError: pass
    return f"{prefix}{max_num + 1:03d}"

def find_user_by_email(email):
    for user in get_users():
        if user.get("email", "").lower() == email.lower():
            return user
    return None

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user"): return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def roles_required(*allowed_roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = session.get("user")
            if not user: return redirect(url_for("login"))
            if user.get("role") not in allowed_roles:
                flash("You do not have access to that page.", "error")
                return redirect(url_for("route_by_role"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator

def can_edit_project(user, project):
    if not user or not project:
        return False
    if user.get("role") == "ADMIN":
        return True
    if user.get("role") == "MANAGER" and project.get("created_by") == user.get("id"):
        return True
    return False

def previous_working_day(date_str):
    current = datetime.strptime(date_str, "%Y-%m-%d").date() - timedelta(days=1)
    while current.weekday() > 4: current -= timedelta(days=1)
    return current.strftime("%Y-%m-%d")

def get_greeting():
    hour = datetime.now().hour
    if hour < 12: return "Good Morning"
    if hour < 17: return "Good Afternoon"
    return "Good Evening"

def derive_billing_rows(staffing_rows):
    projects = {p["id"]: p for p in get_projects()}
    rates = get_billing_rates()
    rate_map = {}
    for r in rates:
        rate_map[(r.get("region", "Global"), r.get("project_type", ""))] = float(r.get("per_fte_rate", 0))
    grouped = {}
    for row in staffing_rows:
        project_id = row.get("project_id", "")
        case_code = row.get("case_code", "")
        project = projects.get(project_id, {})
        project_name = row.get("project_name") or project.get("project_name", "")
        project_type = project.get("project_type", "Other CD/IP Codes")
        region = project.get("region") or project.get("office") or "Global"
        key = (project_id, project_name, project_type, case_code)
        grouped.setdefault(key, {"project_id": project_id, "project_name": project_name, "project_type": project_type, "case_code": case_code, "billable_ftes": 0, "billing_amount": 0, "comments": ""})
        grouped[key]["billable_ftes"] += float(row.get("hours", 0)) / 8.0
        grouped[key]["region"] = region
    output = []
    for item in grouped.values():
        rate = rate_map.get((item.get("region", "Global"), item["project_type"]), 0)
        if rate == 0:
            rate = rate_map.get(("Global", item["project_type"]), 0)
        item["billable_ftes"] = round(item["billable_ftes"], 2)
        item["billing_amount"] = round(item["billable_ftes"] * rate, 2)
        item.pop("region", None)
        output.append(item)
    return output

def build_master_sheet_rows():
    staffing = get_staffing_entries()
    employees = {e["id"]: e for e in get_employees()}
    teams = {t["id"]: t for t in get_teams()}
    projects = {p["id"]: p for p in get_projects()}
    managers = {u["id"]: u for u in get_users()}
    rows = []
    for item in staffing:
        emp = employees.get(item.get("employee_id", ""), {})
        proj = projects.get(item.get("project_id", ""), {})
        mgr = managers.get(item.get("manager_id", ""), {})
        staffing_team_id = resolve_staffing_team_id(item, projects, employees)
        team = teams.get(staffing_team_id, {})
        rows.append({
            "id": item.get("id"),
            "date": item.get("date"),
            "manager_id": item.get("manager_id"),
            "employee_id": item.get("employee_id"),
            "project_id": item.get("project_id"),
            "manager": f'{mgr.get("first_name","")} {mgr.get("last_name","")}'.strip(),
            "employee": emp.get("name", ""),
            "designation": emp.get("designation", ""),
            "team": team.get("team_name", ""),
            "staffing_type": item.get("staffing_type", ""),
            "project": proj.get("project_name", item.get("project_name","")),
            "case_code": item.get("case_code", ""),
            "hours": item.get("hours", 0),
            "comments": item.get("comments", "")
        })
    return rows

def find_by_id(rows, item_id):
    return next((r for r in rows if r.get("id") == item_id), None)

@app.context_processor
def inject_globals():
    return {"current_user": session.get("user"), "dropdowns": get_dropdown_options()}

@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"): return redirect(url_for("route_by_role"))
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        user = find_user_by_email(email)
        if user and user.get("password") == password:
            session["user"] = {"id": user["id"], "email": user["email"], "first_name": user["first_name"], "last_name": user["last_name"], "role": user["role"]}
            return redirect(url_for("route_by_role"))
        flash("Invalid email or password.", "error")
    return render_template("login.html")

@app.route("/route-by-role")
@login_required
def route_by_role():
    role = session["user"]["role"]
    if role == "ADMIN": return redirect(url_for("admin_home"))
    if role == "DIRECTOR": return redirect(url_for("director_home"))
    return redirect(url_for("manager_home"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = session["user"]
    if request.method == "POST":
        users = get_users()
        for item in users:
            if item["id"] == user["id"]:
                item["first_name"] = request.form.get("first_name", item["first_name"]).strip() or item["first_name"]
                item["last_name"] = request.form.get("last_name", item["last_name"]).strip() or item["last_name"]
                new_password = request.form.get("password", "").strip()
                if new_password: item["password"] = new_password
                user["first_name"] = item["first_name"]; user["last_name"] = item["last_name"]; session["user"] = user
                break
        save_users(users); flash("Profile updated.", "success"); return redirect(url_for("profile"))
    return render_template("profile.html")

@app.route("/admin")
@login_required
@roles_required("ADMIN")
def admin_home():
    stats = {"users": len(get_users()), "teams": len(get_teams()), "employees": len(get_employees()), "projects": len(get_projects())}
    return render_template("admin_home.html", stats=stats)

@app.route("/admin/backup-download")
@login_required
@roles_required("ADMIN")
def admin_backup_download():
    mem_zip = BytesIO()
    with zipfile.ZipFile(mem_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for filename in sorted(os.listdir(DATA_DIR)):
            if filename.endswith('.json'):
                filepath = os.path.join(DATA_DIR, filename)
                zf.write(filepath, filename)
    mem_zip.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(mem_zip, mimetype='application/zip', as_attachment=True, download_name=f"bcn_backup_{timestamp}.zip")

@app.route("/admin/master-sheet")
@login_required
@roles_required("ADMIN")
def admin_master_sheet():
    return render_template("admin_master_sheet.html", rows=build_master_sheet_rows())

@app.route("/admin/master-sheet/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_master_sheet_edit(item_id):
    staffing = get_staffing_entries()
    item = find_by_id(staffing, item_id)
    if not item: flash("Master sheet entry not found.", "error"); return redirect(url_for("admin_master_sheet"))
    if request.method == "POST":
        item["date"] = request.form.get("date", "").strip()
        item["manager_id"] = request.form.get("manager_id", "").strip()
        item["employee_id"] = request.form.get("employee_id", "").strip()
        item["project_id"] = request.form.get("project_id", "").strip()
        item["staffing_type"] = request.form.get("staffing_type", "").strip()
        item["case_code"] = request.form.get("case_code", "").strip()
        item["hours"] = float(request.form.get("hours", "0") or 0)
        item["comments"] = request.form.get("comments", "").strip()
        save_staffing_entries(staffing)
        flash("Master sheet entry updated.", "success")
        return redirect(url_for("admin_master_sheet"))
    return render_template("admin_master_sheet_edit.html", item=item, users=get_users(), employees=get_employees(), projects=get_projects(), staffing_options=get_dropdown_options().get("staffing_type", []))

@app.route("/admin/master-sheet/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_master_sheet_delete(item_id):
    staffing = get_staffing_entries()
    target = find_by_id(staffing, item_id)

    if not target:
        flash("Master sheet entry not found.", "error")
        return redirect(url_for("admin_master_sheet"))

    manager_id = target.get("manager_id", "")
    date = target.get("date", "")

    updated_staffing = [r for r in staffing if r.get("id") != item_id]
    save_staffing_entries(updated_staffing)

    # Billing is derived from staffing; after a staffing delete, rebuild billing for that manager/date.
    billing = [
        r for r in get_billing_entries()
        if not (r.get("manager_id") == manager_id and r.get("date") == date)
    ]

    remaining_rows = [
        r for r in updated_staffing
        if r.get("manager_id") == manager_id and r.get("date") == date
    ]
    rebuilt_billing_rows = derive_billing_rows(remaining_rows)

    for row in rebuilt_billing_rows:
        billing.append({
            "id": next_id("BE", billing),
            "date": date,
            "manager_id": manager_id,
            "project_id": row.get("project_id", ""),
            "project_name": row.get("project_name", ""),
            "project_type": row.get("project_type", ""),
            "case_code": row.get("case_code", ""),
            "billable_ftes": float(row.get("billable_ftes", 0) or 0),
            "billing_amount": float(row.get("billing_amount", 0) or 0),
            "comments": "",
        })

    save_billing_entries(billing)
    flash("Master sheet entry deleted. Linked billing has been refreshed.", "success")
    return redirect(url_for("admin_master_sheet"))

@app.route("/admin/export/master-sheet.csv")
@login_required
@roles_required("ADMIN")
def export_master_sheet_csv():
    rows = build_master_sheet_rows()
    output = StringIO(); writer = csv.DictWriter(output, fieldnames=["id","date", "manager", "employee", "designation", "team", "staffing_type", "project", "case_code", "hours", "comments"])
    writer.writeheader(); writer.writerows(rows)
    mem = BytesIO(output.getvalue().encode("utf-8"))
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="master_sheet.csv")

# Keep existing routes behavior for other admin modules from current app by rendering existing templates and saving if present
@app.route("/admin/users")
@login_required
@roles_required("ADMIN")
def admin_users(): return render_template("admin_users.html", rows=get_users())

@app.route("/admin/users/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_users_edit(item_id):
    rows = get_users()
    item = find_by_id(rows, item_id)

    if not item:
        flash("User not found.", "error")
        return redirect(url_for("admin_users"))

    if request.method == "POST":
        item["first_name"] = request.form.get("first_name", "").strip()
        item["last_name"] = request.form.get("last_name", "").strip()
        item["email"] = request.form.get("email", "").strip()
        item["role"] = request.form.get("role", "MANAGER").strip()

        password = request.form.get("password", "").strip()
        if password:
            item["password"] = password

        save_users(rows)
        flash("User updated.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_edit.html", item=item)

@app.route("/admin/users/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_users_delete(item_id):
    save_users([r for r in get_users() if r.get("id") != item_id])
    flash("User deleted.", "success")
    return redirect(url_for("admin_users"))

@app.route("/admin/teams", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_teams():
    rows = get_teams()
    if request.method == "POST":
        approved_headcount_raw = request.form.get("approved_headcount", "").strip()
        approved_annual_budget_raw = request.form.get("approved_annual_budget", "").strip()

        new_row = {
            "id": next_id("T", rows),
            "team_name": request.form.get("team_name", "").strip(),
            "team_type": request.form.getlist("team_type"),
            "team_classification_main": request.form.get("team_classification_main", "").strip(),
            "team_classification_1": request.form.get("team_classification_1", "").strip(),
            "team_classification_2": request.form.get("team_classification_2", "").strip(),
            "manager_id": request.form.get("manager_id", "").strip(),
            "ombudsperson_employee_id": request.form.get("ombudsperson_employee_id", "").strip(),
            "approved_headcount": (safe_float(approved_headcount_raw) if approved_headcount_raw else ""),
            "approved_annual_budget": (safe_float(approved_annual_budget_raw.replace(",", "").replace("$", "")) if approved_annual_budget_raw else ""),
        }

        rows.append(new_row)
        save_teams(rows)
        flash("Team added.", "success")
        return redirect(url_for("admin_teams"))

    return render_template(
        "admin_teams.html",
        rows=rows,
        managers=[u for u in get_users() if u.get("role") == "MANAGER"],
        employees=get_employees(),
    )

@app.route("/admin/teams/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_teams_edit(item_id):
    rows = get_teams()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Team not found.", "error")
        return redirect(url_for("admin_teams"))

    if request.method == "POST":
        item["team_name"] = request.form.get("team_name", "").strip()
        item["team_type"] = request.form.getlist("team_type")
        item["team_classification_main"] = request.form.get("team_classification_main", "").strip()
        item["team_classification_1"] = request.form.get("team_classification_1", "").strip()
        item["team_classification_2"] = request.form.get("team_classification_2", "").strip()
        item["manager_id"] = request.form.get("manager_id", "").strip()
        item["ombudsperson_employee_id"] = request.form.get("ombudsperson_employee_id", "").strip()

        approved_headcount_raw = request.form.get("approved_headcount", "").strip()
        approved_annual_budget_raw = request.form.get("approved_annual_budget", "").strip()
        item["approved_headcount"] = (safe_float(approved_headcount_raw) if approved_headcount_raw else "")
        item["approved_annual_budget"] = (safe_float(approved_annual_budget_raw.replace(",", "").replace("$", "")) if approved_annual_budget_raw else "")

        save_teams(rows)
        flash("Team updated.", "success")
        return redirect(url_for("admin_teams"))

    return render_template(
        "admin_team_edit.html",
        item=item,
        managers=[u for u in get_users() if u.get("role") == "MANAGER"],
        employees=get_employees()
    )

@app.route("/admin/teams/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_teams_delete(item_id):
    save_teams([r for r in get_teams() if r.get("id") != item_id])
    flash("Team deleted.", "success")
    return redirect(url_for("admin_teams"))

@app.route("/admin/employees")
@login_required
@roles_required("ADMIN")
def admin_employees(): return render_template("admin_employees.html", rows=get_employees(), teams=get_teams())

@app.route("/admin/employees/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_employees_edit(item_id):
    rows = get_employees()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Employee not found.", "error")
        return redirect(url_for("admin_employees"))

    if request.method == "POST":
        item["name"] = request.form.get("name", "").strip()
        item["employee_code"] = request.form.get("employee_code", "").strip()
        item["gender"] = request.form.get("gender", "").strip()
        item["designation"] = request.form.get("designation", "").strip()
        item["team_id"] = request.form.get("team_id", "").strip()

        save_employees(rows)
        flash("Employee updated.", "success")
        return redirect(url_for("admin_employees"))

    return render_template("admin_employee_edit.html", item=item, teams=get_teams())

@app.route("/admin/employees/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_employees_delete(item_id):
    save_employees([r for r in get_employees() if r.get("id") != item_id])
    flash("Employee deleted.", "success")
    return redirect(url_for("admin_employees"))

@app.route("/admin/projects", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_projects():
    rows = get_projects()
    if request.method == "POST":
        new_row = {"id": next_id("P", rows)}
        for field in FULL_PROJECT_FIELDS:
            new_row[field] = request.form.get(field, "").strip()
        new_row["team_id"] = request.form.get("team_id", "").strip()
        rows.append(new_row)
        save_projects(rows)
        flash("Project added.", "success")
        return redirect(url_for("admin_projects"))
    return render_template("admin_projects.html", rows=rows, teams=get_teams(), users=get_users(), dropdowns=get_dropdown_options())

@app.route("/admin/projects/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_projects_edit(item_id):
    rows = get_projects()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Project not found.", "error")
        return redirect(url_for("admin_projects"))

    if request.method == "POST":
        for key in FULL_PROJECT_FIELDS:
            item[key] = request.form.get(key, "").strip()

        save_projects(rows)
        flash("Project updated.", "success")
        return redirect(url_for("admin_projects"))

    return render_template(
        "admin_project_edit.html",
        item=item,
        teams=get_teams(),
        users=get_users()
    )

@app.route("/admin/projects/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_projects_delete(item_id):
    save_projects([r for r in get_projects() if r.get("id") != item_id])
    flash("Project deleted.", "success")
    return redirect(url_for("admin_projects"))

@app.route("/admin/billing-rates", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_billing_rates():
    rows = get_billing_rates()
    if request.method == "POST":
        new_row = {
            "id": next_id("BR", rows),
            "region": request.form.get("region", "").strip(),
            "project_type": request.form.get("project_type", "").strip(),
            "per_fte_rate": request.form.get("per_fte_rate", "").strip() or "0",
        }
        rows.append(new_row)
        save_billing_rates(rows)
        flash("Billing rate added.", "success")
        return redirect(url_for("admin_billing_rates"))
    return render_template("admin_billing_rates.html", rows=rows, dropdowns=get_dropdown_options())

@app.route("/admin/billing-rates/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_billing_rates_edit(item_id):
    rows = get_billing_rates()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Billing rate not found.", "error")
        return redirect(url_for("admin_billing_rates"))

    if request.method == "POST":
        item["region"] = request.form.get("region", "").strip()
        item["project_type"] = request.form.get("project_type", "").strip()
        item["per_fte_rate"] = request.form.get("per_fte_rate", "").strip() or "0"

        save_billing_rates(rows)
        flash("Billing rate updated.", "success")
        return redirect(url_for("admin_billing_rates"))

    return render_template("admin_billing_rate_edit.html", item=item)

@app.route("/admin/billing-rates/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_billing_rates_delete(item_id):
    save_billing_rates([r for r in get_billing_rates() if r.get("id") != item_id])
    flash("Billing rate deleted.", "success")
    return redirect(url_for("admin_billing_rates"))

@app.route("/admin/cost-rates", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_cost_rates():
    rows = get_cost_rates()
    if request.method == "POST":
        new_row = {
            "id": next_id("CR", rows),
            "designation": request.form.get("designation", "").strip(),
            "per_fte_rate": request.form.get("per_fte_rate", "").strip() or "0",
        }
        rows.append(new_row)
        save_cost_rates(rows)
        flash("Cost rate added.", "success")
        return redirect(url_for("admin_cost_rates"))
    return render_template("admin_cost_rates.html", rows=rows, dropdowns=get_dropdown_options())

@app.route("/admin/cost-rates/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("ADMIN")
def admin_cost_rates_edit(item_id):
    rows = get_cost_rates()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Cost rate not found.", "error")
        return redirect(url_for("admin_cost_rates"))

    if request.method == "POST":
        item["designation"] = request.form.get("designation", "").strip()
        item["per_fte_rate"] = request.form.get("per_fte_rate", "").strip() or "0"

        save_cost_rates(rows)
        flash("Cost rate updated.", "success")
        return redirect(url_for("admin_cost_rates"))

    return render_template("admin_cost_rate_edit.html", item=item)

@app.route("/admin/cost-rates/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_cost_rates_delete(item_id):
    save_cost_rates([r for r in get_cost_rates() if r.get("id") != item_id])
    flash("Cost rate deleted.", "success")
    return redirect(url_for("admin_cost_rates"))

@app.route("/admin/reports")
@login_required
@roles_required("ADMIN")
def admin_reports(): return render_template("admin_reports.html", rows=get_reports())

@app.route("/admin/reports/delete/<item_id>")
@login_required
@roles_required("ADMIN")
def admin_reports_delete(item_id):
    save_reports([r for r in get_reports() if r.get("id") != item_id])
    flash("Report deleted.", "success")
    return redirect(url_for("admin_reports"))

@app.route("/admin/dropdown-options")
@login_required
@roles_required("ADMIN")
def admin_dropdown_options(): return render_template("admin_dropdown_options.html", options=get_dropdown_options())

@app.route("/admin/dropdown-options/delete", methods=["POST"])
@login_required
@roles_required("ADMIN")
def admin_dropdown_options_delete():
    options = get_dropdown_options()
    key = request.form.get("key", "").strip()
    value = request.form.get("value", "").strip()

    if key in options:
        options[key] = [v for v in options[key] if v != value]
        save_dropdown_options(options)
        flash("Option removed.", "success")

    return redirect(url_for("admin_dropdown_options"))

@app.route("/reports/download/<item_id>")
@login_required
def reports_download(item_id):
    item = next((r for r in get_reports() if r.get("id") == item_id), None)
    if not item: flash("Report not found.", "error"); return redirect(url_for("route_by_role"))
    stored = item.get("stored_file", "")
    if stored and os.path.exists(os.path.join(UPLOAD_DIR, stored)):
        return send_file(os.path.join(UPLOAD_DIR, stored), as_attachment=True, download_name=item.get("file_name") or stored)
    flash("Physical file not uploaded yet. This is still a placeholder metadata record.", "error")
    return redirect(request.referrer or url_for("route_by_role"))

def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return None

def month_label(dt):
    return dt.strftime("%b %Y")

def money_fmt(value):
    return f"${value:,.0f}"

def pct_fmt(value):
    return f"{value:.0f}%"

def safe_float(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def working_days_in_month(year, month):
    if month == 12:
        next_month = datetime(year + 1, 1, 1).date()
    else:
        next_month = datetime(year, month + 1, 1).date()
    start = datetime(year, month, 1).date()
    days = 0
    current = start
    while current < next_month:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days

def ordinal_day(n):
    if 10 <= (n % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def format_compact_date(dt):
    return f"{ordinal_day(dt.day)} {dt.strftime('%b')}"

def format_date_ranges(date_values):
    if not date_values:
        return "-"
    ordered = sorted(set(date_values))
    ranges = []
    start = ordered[0]
    end = ordered[0]
    for dt in ordered[1:]:
        if (dt - end).days == 1:
            end = dt
            continue
        ranges.append((start, end))
        start = dt
        end = dt
    ranges.append((start, end))

    labels = []
    for r_start, r_end in ranges:
        if r_start == r_end:
            labels.append(format_compact_date(r_start))
        else:
            labels.append(f"{format_compact_date(r_start)}-{format_compact_date(r_end)}")
    return ", ".join(labels)

def get_available_reporting_months():
    staffing_entries = get_staffing_entries()
    billing_entries = get_billing_entries()

    month_set = set()

    for row in staffing_entries:
        dt = parse_date(row.get("date", ""))
        if dt:
            month_set.add((dt.year, dt.month))

    for row in billing_entries:
        dt = parse_date(row.get("date", ""))
        if dt:
            month_set.add((dt.year, dt.month))

    if not month_set:
        today = datetime.now().date()
        month_set.add((today.year, today.month))

    ordered = sorted(month_set)
    return [datetime(year, month, 1).strftime("%b %Y") for year, month in ordered]

def parse_selected_month(selected_month):
    try:
        dt = datetime.strptime(selected_month, "%b %Y")
        return dt.year, dt.month
    except Exception:
        today = datetime.now().date()
        return today.year, today.month

def get_reporting_dates():
    staffing_entries = get_staffing_entries()
    billing_entries = get_billing_entries()
    dates = set()

    for row in staffing_entries + billing_entries:
        dt = parse_date(row.get("date", ""))
        if dt:
            dates.add(dt)

    if not dates:
        dates.add(datetime.now().date())

    return sorted(dates)

def get_director_timeline_options(period_scope):
    period_scope = (period_scope or "monthly").strip().lower()
    if period_scope not in ["weekly", "monthly", "ytd"]:
        period_scope = "monthly"

    dates = get_reporting_dates()

    if period_scope == "weekly":
        week_starts = sorted({d - timedelta(days=d.weekday()) for d in dates})
        return [
            {
                "value": ws.strftime("%Y-%m-%d"),
                "label": f"{ws.strftime('%d %b %Y')} - {(ws + timedelta(days=6)).strftime('%d %b %Y')}"
            }
            for ws in week_starts
        ]

    if period_scope == "ytd":
        years = sorted({d.year for d in dates})
        return [{"value": str(y), "label": str(y)} for y in years]

    month_set = sorted({(d.year, d.month) for d in dates})
    return [
        {
            "value": datetime(y, m, 1).strftime("%b %Y"),
            "label": datetime(y, m, 1).strftime("%b %Y")
        }
        for y, m in month_set
    ]

def build_director_gantt(active_view="weekly", anchor_date=None, manager_filter=None, 
                         team_class_main_filter=None, team_class_1_filter=None, team_class_2_filter=None):
    """Build gantt data for director project view with filtering support."""
    if anchor_date is None:
        anchor_date = datetime.now().date()
    
    employees = get_employees()
    teams = get_teams()
    users = get_users()
    projects = get_projects()
    staffing_entries = get_staffing_entries()
    billing_entries = get_billing_entries()
    
    project_map = {p["id"]: p for p in projects if p.get("id")}
    team_map = {t["id"]: t for t in teams if t.get("id")}
    user_map = {u["id"]: u for u in users if u.get("id")}
    employee_map = {e["id"]: e for e in employees if e.get("id")}
    
    # Filter teams based on director filters
    filtered_teams = []
    for team in teams:
        # Filter by manager
        if manager_filter and manager_filter != "All":
            if team.get("manager_id") != manager_filter:
                continue
        
        # Filter by team classification main
        if team_class_main_filter and team_class_main_filter != "All":
            if team.get("team_classification_main") != team_class_main_filter:
                continue
        
        # Filter by team classification 1
        if team_class_1_filter and team_class_1_filter != "All":
            if team.get("team_classification_1") != team_class_1_filter:
                continue
        
        # Filter by team classification 2
        if team_class_2_filter and team_class_2_filter != "All":
            if team.get("team_classification_2") != team_class_2_filter:
                continue
        
        filtered_teams.append(team)
    
    # Determine date range based on view
    today = datetime.now().date()
    if active_view == "weekly":
        period_start = anchor_date - timedelta(days=anchor_date.weekday())
        visible_dates = [period_start + timedelta(days=i) for i in range(5)]
        prev_anchor = period_start - timedelta(days=7)
        next_anchor = period_start + timedelta(days=7)
        period_label = f"{visible_dates[0].strftime('%d %b %Y')} - {visible_dates[-1].strftime('%d %b %Y')}"
    else:  # monthly
        period_start = anchor_date.replace(day=1)
        if period_start.month == 12:
            next_month_start = period_start.replace(year=period_start.year + 1, month=1, day=1)
        else:
            next_month_start = period_start.replace(month=period_start.month + 1, day=1)
        
        period_end = next_month_start - timedelta(days=1)
        visible_dates = []
        cursor = period_start
        while cursor <= period_end:
            if cursor.weekday() < 5:
                visible_dates.append(cursor)
            cursor += timedelta(days=1)
        
        prev_anchor = (period_start - timedelta(days=1)).replace(day=1)
        next_anchor = next_month_start
        period_label = period_start.strftime("%B %Y")
    
    visible_date_strings = [d.strftime("%Y-%m-%d") for d in visible_dates]
    week_date_labels = [d.strftime("%d %b") for d in visible_dates]
    week_day_labels = [d.strftime("%a").upper() for d in visible_dates]
    
    # Collect staffing entries in the visible range for filtered teams (team resolved by project first)
    filtered_team_ids = {t["id"] for t in filtered_teams}
    entries_in_range = [
        s for s in staffing_entries
        if s.get("date") in visible_date_strings and resolve_staffing_team_id(s, project_map, employee_map) in filtered_team_ids
    ]

    billing_in_range = []
    for b in billing_entries:
        if b.get("date") not in visible_date_strings:
            continue
        project_id = b.get("project_id", "")
        project = project_map.get(project_id, {})
        if project.get("team_id") not in filtered_team_ids:
            continue
        billing_in_range.append(b)
    
    # Build gantt rows for each team
    gantt_rows = []
    for team in filtered_teams:
        team_entries = [s for s in entries_in_range if resolve_staffing_team_id(s, project_map, employee_map) == team["id"]]
        
        if not team_entries:
            gantt_rows.append({
                "team_name": team.get("team_name", ""),
                "bars": [],
                "is_empty": True
            })
            continue
        
        # Collect project data with day-level FTEs + totals for tooltip
        project_day_ftes = {}   # {project_key: {date_str: fte_on_day}}
        project_info = {}       # {project_key: metadata}

        def resolve_project_meta(project_id, fallback_name=""):
            project = project_map.get(project_id, {}) if project_id else {}
            full_name = project.get("project_name", "") or fallback_name or "Unspecified"
            region = (
                project.get("billing_region", "")
                or project.get("office", "")
                or project.get("case_delivery_primary_location", "")
                or "Global"
            )
            return {
                "name": full_name,
                "region": region,
                "requestor": project.get("requestor", "") or "-",
            }
        
        for s in team_entries:
            if s.get("project_id") and s.get("project_id") in project_map:
                project_key = s["project_id"]
                label = project_map[s["project_id"]].get("project_name", "")
            elif s.get("project_name"):
                project_key = s.get("project_name")
                label = s.get("project_name", "")
            else:
                project_key = s.get("staffing_type", "Unspecified")
                label = s.get("staffing_type", "Unspecified")
            
            if not label:
                continue
            
            date_key = s.get("date")
            hours = float(s.get("hours", 0) or 0)
            project_day_ftes.setdefault(project_key, {})
            project_day_ftes[project_key][date_key] = project_day_ftes[project_key].get(date_key, 0.0) + (hours / 8.0)
            if project_key not in project_info:
                meta = resolve_project_meta(s.get("project_id", ""), s.get("project_name", "") or label)
                project_info[project_key] = {
                    "name": meta["name"],
                    "region": meta["region"],
                    "requestor": meta["requestor"],
                    "total_hours": 0.0,
                    "total_billing": 0.0,
                }
            project_info[project_key]["total_hours"] += hours

        for b in billing_in_range:
            if b.get("project_id") and b.get("project_id") in project_map:
                project_key = b.get("project_id")
            elif b.get("project_name"):
                project_key = b.get("project_name")
            else:
                continue
            if project_key in project_info:
                project_info[project_key]["total_billing"] += float(b.get("billing_amount", 0) or 0)
        
        # Build occupied-day indexes for overlap detection
        project_indexes = {}
        for project_key, day_map in project_day_ftes.items():
            indexes_in_range = {
                visible_date_strings.index(d)
                for d, fte in day_map.items()
                if d in visible_date_strings and fte > 0
            }
            if indexes_in_range:
                project_indexes[project_key] = indexes_in_range
        
        # Detect overlapping projects and assign to levels
        def projects_overlap(idx_set_1, idx_set_2):
            return len(idx_set_1.intersection(idx_set_2)) > 0
        
        project_to_level = {}
        sorted_projects = sorted(project_indexes.items(), key=lambda x: min(x[1]))
        
        for project_key, idx_set in sorted_projects:
            level = 0
            for other_project_key, other_level in project_to_level.items():
                if other_level == level and projects_overlap(idx_set, project_indexes[other_project_key]):
                    level += 1
            project_to_level[project_key] = level
        
        # Build bars
        bars = []
        for project_key in sorted(project_indexes.keys()):
            project_name = project_info[project_key]["name"]
            level = project_to_level[project_key]
            
            day_fte_by_idx = {}
            for d, fte in project_day_ftes.get(project_key, {}).items():
                if d in visible_date_strings and fte > 0:
                    day_fte_by_idx[visible_date_strings.index(d)] = fte

            dates_for_project = sorted(day_fte_by_idx.keys())
            
            segments = []
            if dates_for_project:
                i = 0
                while i < len(dates_for_project):
                    start_idx = dates_for_project[i]
                    end_idx = start_idx
                    current_fte = day_fte_by_idx[start_idx]
                    
                    while (
                        i + 1 < len(dates_for_project)
                        and dates_for_project[i + 1] == dates_for_project[i] + 1
                        and abs(day_fte_by_idx[dates_for_project[i + 1]] - current_fte) < 1e-9
                    ):
                        i += 1
                        end_idx = dates_for_project[i]
                    
                    span = end_idx - start_idx + 1
                    fte_label = f"{current_fte:.1f}".rstrip("0").rstrip(".")
                    label = f"{project_name} - {fte_label} FTEs"
                    segments.append({
                        "style": f"grid-column: {start_idx + 1} / span {span};",
                        "label": label
                    })
                    i += 1
            
            # Calculate daily FTE and total days from visible data
            num_days = len(day_fte_by_idx)
            if num_days > 0:
                # Average FTE across visible days
                daily_fte = round(sum(day_fte_by_idx.values()) / num_days, 2)
            else:
                daily_fte = 0.0
            
            bars.append({
                "project_key": project_key,
                "project_name": project_name,
                "level": level,
                "tooltip": {
                    "project_name": project_info[project_key].get("name", project_name),
                    "daily_ftes": daily_fte,
                    "total_days": num_days,
                    "region": project_info[project_key].get("region", "Global"),
                    "requestor": project_info[project_key].get("requestor", "-"),
                    "total_billing": f"${project_info[project_key].get('total_billing', 0.0):,.2f}",
                },
                "segments": segments
            })
        
        max_level = max([b["level"] for b in bars]) if bars else 0
        gantt_rows.append({
            "team_name": team.get("team_name", ""),
            "bars": bars,
            "is_empty": False,
            "max_level": max_level
        })
    
    return {
        "gantt_rows": gantt_rows,
        "visible_dates": visible_date_strings,
        "week_date_labels": week_date_labels,
        "week_day_labels": week_day_labels,
        "period_label": period_label,
        "prev_anchor": prev_anchor.strftime("%Y-%m-%d"),
        "next_anchor": next_anchor.strftime("%Y-%m-%d"),
        "current_anchor": anchor_date.strftime("%Y-%m-%d")
    }

def build_director_metrics(selected_timeline, period_scope="monthly", tracker_summary="auto"):
    year = datetime.now().year
    month = datetime.now().month
    selected_week_start = None
    period_scope = (period_scope or "monthly").strip().lower()
    if period_scope not in ["weekly", "monthly", "ytd"]:
        period_scope = "monthly"

    selected_period_label = ""

    if period_scope == "monthly":
        year, month = parse_selected_month(selected_timeline)
        selected_period_label = selected_timeline
    elif period_scope == "weekly":
        try:
            week_start = datetime.strptime(selected_timeline, "%Y-%m-%d").date()
        except Exception:
            week_start = datetime.now().date() - timedelta(days=datetime.now().date().weekday())
        selected_week_start = week_start
        week_end = week_start + timedelta(days=6)
        year, month = week_start.year, week_start.month
        selected_period_label = f"{week_start.strftime('%d %b %Y')} - {week_end.strftime('%d %b %Y')}"
    else:
        try:
            year = int(selected_timeline)
        except Exception:
            year = datetime.now().year
        all_dates = get_reporting_dates()
        months_for_year = sorted({d.month for d in all_dates if d.year == year})
        month = months_for_year[-1] if months_for_year else datetime.now().month
        selected_period_label = str(year)

    tracker_summary = (tracker_summary or "auto").strip().lower()
    if tracker_summary not in ["auto", "project", "week", "month", "year", "type_of_project", "team_classification_main", "team_classification_1", "team_classification_2", "region"]:
        tracker_summary = "auto"

    effective_tracker_summary = tracker_summary
    if tracker_summary == "auto":
        if period_scope == "weekly":
            effective_tracker_summary = "week"
        elif period_scope == "monthly":
            effective_tracker_summary = "month"
        else:
            effective_tracker_summary = "year"

    employees = get_employees()
    teams = get_teams()
    users = get_users()
    projects = get_projects()
    staffing_entries = get_staffing_entries()
    billing_entries = get_billing_entries()
    cost_rates = get_cost_rates()

    project_map = {p["id"]: p for p in projects if p.get("id")}
    team_map = {t["id"]: t for t in teams if t.get("id")}
    user_map = {u["id"]: u for u in users if u.get("id")}
    employee_map = {e["id"]: e for e in employees if e.get("id")}

    cost_rate_map = {}
    for row in cost_rates:
        designation = (row.get("designation") or "").strip()
        cost_rate_map[designation] = safe_float(row.get("per_fte_rate", 0))

    def rows_for_scope(rows, scope, y, m, week_anchor=None):
        parsed = []
        for r in rows:
            dt = parse_date(r.get("date", ""))
            if dt:
                parsed.append((r, dt))

        if scope == "monthly":
            return [r for r, dt in parsed if dt.year == y and dt.month == m]

        if scope == "ytd":
            return [r for r, dt in parsed if dt.year == y and dt.month <= m]

        # weekly: prefer explicit selected week anchor when provided
        if week_anchor:
            anchor_dt = week_anchor
        else:
            # fallback to the last dated entry in selected month (or month end)
            selected_month_dates = [dt for _, dt in parsed if dt.year == y and dt.month == m]
            if selected_month_dates:
                anchor_dt = max(selected_month_dates)
            else:
                if m == 12:
                    anchor_dt = datetime(y, 12, 31).date()
                else:
                    anchor_dt = (datetime(y, m + 1, 1).date() - timedelta(days=1))

        week_start = anchor_dt - timedelta(days=anchor_dt.weekday())
        week_end = week_start + timedelta(days=6)
        return [r for r, dt in parsed if week_start <= dt <= week_end]

    month_staffing = rows_for_scope(staffing_entries, period_scope, year, month, selected_week_start)
    month_billing = rows_for_scope(billing_entries, period_scope, year, month, selected_week_start)

    if period_scope == "monthly":
        if month == 1:
            prev_year, prev_month = year - 1, 12
        else:
            prev_year, prev_month = year, month - 1
        prev_month_staffing = rows_for_scope(staffing_entries, period_scope, prev_year, prev_month)
        prev_month_billing = rows_for_scope(billing_entries, period_scope, prev_year, prev_month)
    elif period_scope == "weekly":
        # Previous comparable week relative to selected week
        current_week_start = selected_week_start or (datetime.now().date() - timedelta(days=datetime.now().date().weekday()))
        prev_week_start = current_week_start - timedelta(days=7)
        prev_month_staffing = rows_for_scope(staffing_entries, "weekly", prev_week_start.year, prev_week_start.month, prev_week_start)
        prev_month_billing = rows_for_scope(billing_entries, "weekly", prev_week_start.year, prev_week_start.month, prev_week_start)
    else:
        # ytd compares against prior-year YTD through the same month
        prev_month_staffing = rows_for_scope(staffing_entries, "ytd", year - 1, month)
        prev_month_billing = rows_for_scope(billing_entries, "ytd", year - 1, month)

    # FTEs by average daily staffed hours in month
    current_days = sorted({r.get("date") for r in month_staffing if r.get("date")})
    prev_days = sorted({r.get("date") for r in prev_month_staffing if r.get("date")})

    total_ftes = 0.0
    if current_days:
        total_ftes = sum(
            sum(safe_float(r.get("hours", 0)) for r in month_staffing if r.get("date") == day) / 8.0
            for day in current_days
        ) / len(current_days)

    prev_total_ftes = 0.0
    if prev_days:
        prev_total_ftes = sum(
            sum(safe_float(r.get("hours", 0)) for r in prev_month_staffing if r.get("date") == day) / 8.0
            for day in prev_days
        ) / len(prev_days)

    fte_trend = f"{total_ftes - prev_total_ftes:+.1f}"

    total_billing = sum(safe_float(r.get("billing_amount", 0)) for r in month_billing)
    prev_total_billing = sum(safe_float(r.get("billing_amount", 0)) for r in prev_month_billing)
    billing_trend = money_fmt(total_billing - prev_total_billing)

    # Cost / recovery
    # Gross cost based on staffed hours and cost rate per designation
    gross_cost = 0.0
    for row in month_staffing:
        emp = employee_map.get(row.get("employee_id", ""), {})
        designation = emp.get("designation", "")
        per_fte_rate = cost_rate_map.get(designation, 0.0)
        gross_cost += (safe_float(row.get("hours", 0)) / 8.0) * per_fte_rate

    recovery_pct = (total_billing / gross_cost * 100.0) if gross_cost > 0 else 0.0

    # Gender ratio
    male_count = len([e for e in employees if (e.get("gender") or "").strip().lower() == "male"])
    female_count = len([e for e in employees if (e.get("gender") or "").strip().lower() == "female"])
    gender_ratio = f"{male_count}:{female_count}" if (male_count or female_count) else "-"
    total_gender_known = male_count + female_count
    female_share = pct_fmt((female_count / total_gender_known) * 100.0) if total_gender_known else "-"

    # NPS / CTSU currently unavailable in your live data model
    nps_value = "-"
    ctsu_value = "-"
    nps_trend = "-"
    ctsu_trend = "-"

    # Major clients by billing in selected month
    client_billing = {}
    client_projects = {}

    for row in month_billing:
        project = project_map.get(row.get("project_id", ""), {})
        client_name = (project.get("client_name") or "Unknown Client").strip() or "Unknown Client"
        project_name = project.get("project_name", row.get("project_name", ""))
        client_billing[client_name] = client_billing.get(client_name, 0.0) + safe_float(row.get("billing_amount", 0))
        client_projects.setdefault(client_name, set()).add(project_name)

    major_clients = []
    for client_name, amount in sorted(client_billing.items(), key=lambda x: x[1], reverse=True)[:5]:
        major_clients.append({
            "name": client_name,
            "projects": len([p for p in client_projects.get(client_name, set()) if p]),
            "billing": money_fmt(amount)
        })

    # Notes from live data
    notes = []
    notes.append(f"Billing for {selected_period_label} is {money_fmt(total_billing)}.")
    notes.append(f"Average staffed FTEs for {selected_period_label} is {total_ftes:.1f}.")
    notes.append(f"Recovery for {selected_period_label} is {pct_fmt(recovery_pct)} based on available cost rates.")
    if major_clients:
        notes.append(f"Top billed client this month is {major_clients[0]['name']} at {major_clients[0]['billing']}.")
    else:
        notes.append("No billed client work found for the selected month.")

    # Monthly trend across all available months
    available_months = get_available_reporting_months()
    monthly_trend = []

    for month_name in available_months:
        y, m = parse_selected_month(month_name)
        trend_staffing = [r for r in staffing_entries if (parse_date(r.get("date", "")) and parse_date(r.get("date", "")).year == y and parse_date(r.get("date", "")).month == m)]
        trend_billing = [r for r in billing_entries if (parse_date(r.get("date", "")) and parse_date(r.get("date", "")).year == y and parse_date(r.get("date", "")).month == m)]

        trend_days = sorted({r.get("date") for r in trend_staffing if r.get("date")})
        trend_ftes = 0.0
        if trend_days:
            trend_ftes = sum(
                sum(safe_float(r.get("hours", 0)) for r in trend_staffing if r.get("date") == day) / 8.0
                for day in trend_days
            ) / len(trend_days)

        trend_billing_total = sum(safe_float(r.get("billing_amount", 0)) for r in trend_billing)

        trend_cost = 0.0
        for row in trend_staffing:
            emp = employee_map.get(row.get("employee_id", ""), {})
            designation = emp.get("designation", "")
            per_fte_rate = cost_rate_map.get(designation, 0.0)
            trend_cost += (safe_float(row.get("hours", 0)) / 8.0) * per_fte_rate

        trend_recovery = (trend_billing_total / trend_cost * 100.0) if trend_cost > 0 else 0.0

        monthly_trend.append({
            "month": month_name,
            "ftes": f"{trend_ftes:.1f}",
            "billing": money_fmt(trend_billing_total),
            "recovery": pct_fmt(trend_recovery),
            "nps": "-",
            "ctsu": "-"
        })

    # Billing Summary
    gs_team_ids = {t["id"] for t in teams if t.get("team_classification_main", "").strip().lower() == "global sustainability"}
    sr_team_ids = {t["id"] for t in teams if t.get("team_classification_main", "").strip().lower() == "s&r"}

    def month_rows_for_team_ids(team_ids_subset):
        rows = []
        for month_name in available_months:
            y, m = parse_selected_month(month_name)

            m_staff = [
                r for r in staffing_entries
                if (
                    parse_date(r.get("date", "")) and
                    parse_date(r.get("date", "")).year == y and
                    parse_date(r.get("date", "")).month == m and
                    resolve_staffing_team_id(r, project_map, employee_map) in team_ids_subset
                )
            ]

            m_bill = [
                r for r in billing_entries
                if (
                    parse_date(r.get("date", "")) and
                    parse_date(r.get("date", "")).year == y and
                    parse_date(r.get("date", "")).month == m and
                    project_map.get(r.get("project_id", ""), {}).get("team_id", "") in team_ids_subset
                )
            ]

            m_days = sorted({r.get("date") for r in m_staff if r.get("date")})
            avg_ftes = 0.0
            if m_days:
                avg_ftes = sum(
                    sum(safe_float(r.get("hours", 0)) for r in m_staff if r.get("date") == day) / 8.0
                    for day in m_days
                ) / len(m_days)

            rows.append({
                "month": month_name,
                "gs_headcount": f"{avg_ftes:.1f}",
                "billing": money_fmt(sum(safe_float(r.get("billing_amount", 0)) for r in m_bill))
            })
        return rows

    global_sustainability_rows = month_rows_for_team_ids(gs_team_ids)

    # S&R style table
    sr_rows = []
    for month_name in available_months:
        y, m = parse_selected_month(month_name)

        sr_staff = [
            r for r in staffing_entries
            if (
                parse_date(r.get("date", "")) and
                parse_date(r.get("date", "")).year == y and
                parse_date(r.get("date", "")).month == m and
                resolve_staffing_team_id(r, project_map, employee_map) in sr_team_ids
            )
        ]

        sr_bill = [
            r for r in billing_entries
            if (
                parse_date(r.get("date", "")) and
                parse_date(r.get("date", "")).year == y and
                parse_date(r.get("date", "")).month == m and
                project_map.get(r.get("project_id", ""), {}).get("team_id", "") in sr_team_ids
            )
        ]

        sr_days = sorted({r.get("date") for r in sr_staff if r.get("date")})
        headcount = 0.0
        if sr_days:
            headcount = sum(
                sum(safe_float(r.get("hours", 0)) for r in sr_staff if r.get("date") == day) / 8.0
                for day in sr_days
            ) / len(sr_days)

        gross_cost_sr = 0.0
        for row in sr_staff:
            emp = employee_map.get(row.get("employee_id", ""), {})
            designation = emp.get("designation", "")
            per_fte_rate = cost_rate_map.get(designation, 0.0)
            gross_cost_sr += (safe_float(row.get("hours", 0)) / 8.0) * per_fte_rate

        total_recovery_sr = sum(safe_float(r.get("billing_amount", 0)) for r in sr_bill)
        recovery_pct_sr = (total_recovery_sr / gross_cost_sr * 100.0) if gross_cost_sr > 0 else 0.0
        client_work = total_recovery_sr
        internal_other = max(gross_cost_sr - total_recovery_sr, 0.0)
        working_days = working_days_in_month(y, m)

        sr_rows.append({
            "month": month_name,
            "headcount": f"{headcount:.1f}",
            "monthly_budget": money_fmt(gross_cost_sr),
            "working_days": working_days,
            "ftes_based_on_budget": f"{headcount:.1f}",
            "excess_ftes": "0.0",
            "cost_of_excess_ftes": money_fmt(0),
            "gross_cost": money_fmt(gross_cost_sr),
            "minimum_recovery_needed": money_fmt(gross_cost_sr),
            "retainer": money_fmt(0),
            "total_recovery": money_fmt(total_recovery_sr),
            "client_work": money_fmt(client_work),
            "internal_other": money_fmt(internal_other),
            "recovery_pct": pct_fmt(recovery_pct_sr),
            "client_recovery_pct": pct_fmt(recovery_pct_sr),
            "other_recovery_pct": pct_fmt(0),
            "net_cost_to_practice": money_fmt(max(gross_cost_sr - total_recovery_sr, 0.0)),
            "budget_variance": money_fmt(total_recovery_sr - gross_cost_sr)
        })

    def normalized_num(value):
        if isinstance(value, str):
            value = value.replace(",", "").replace("$", "").strip()
        return safe_float(value)

    def approved_headcount_display(team_ids_subset):
        scoped_teams = [t for t in teams if t.get("id") in team_ids_subset]
        manual_total = 0.0
        has_manual = False
        for t in scoped_teams:
            raw = t.get("approved_headcount", "")
            if str(raw).strip() != "":
                manual_total += normalized_num(raw)
                has_manual = True
        if has_manual:
            return (f"{manual_total:.1f}".rstrip("0").rstrip("."))
        return f"{len([e for e in employees if e.get('team_id') in team_ids_subset])}"

    def approved_annual_budget_display(team_ids_subset):
        scoped_teams = [t for t in teams if t.get("id") in team_ids_subset]
        manual_total = 0.0
        has_manual = False
        for t in scoped_teams:
            raw = t.get("approved_annual_budget", "")
            if str(raw).strip() != "":
                manual_total += normalized_num(raw)
                has_manual = True
        if has_manual:
            return money_fmt(manual_total)
        auto_total = sum(cost_rate_map.get(e.get("designation", ""), 0.0) * 12 for e in employees if e.get("team_id") in team_ids_subset)
        return money_fmt(auto_total)

    # Martha summary from all teams
    martha_rows = []
    for month_name in available_months:
        y, m = parse_selected_month(month_name)

        m_staff = [
            r for r in staffing_entries
            if (
                parse_date(r.get("date", "")) and
                parse_date(r.get("date", "")).year == y and
                parse_date(r.get("date", "")).month == m
            )
        ]
        m_bill = [
            r for r in billing_entries
            if (
                parse_date(r.get("date", "")) and
                parse_date(r.get("date", "")).year == y and
                parse_date(r.get("date", "")).month == m
            )
        ]

        m_days = sorted({r.get("date") for r in m_staff if r.get("date")})
        avg_ftes = 0.0
        if m_days:
            avg_ftes = sum(
                sum(safe_float(r.get("hours", 0)) for r in m_staff if r.get("date") == day) / 8.0
                for day in m_days
            ) / len(m_days)

        gross_cost_all = 0.0
        for row in m_staff:
            emp = employee_map.get(row.get("employee_id", ""), {})
            designation = emp.get("designation", "")
            per_fte_rate = cost_rate_map.get(designation, 0.0)
            gross_cost_all += (safe_float(row.get("hours", 0)) / 8.0) * per_fte_rate

        billed_recovery = sum(safe_float(r.get("billing_amount", 0)) for r in m_bill)
        recovery_total_pct = (billed_recovery / gross_cost_all * 100.0) if gross_cost_all > 0 else 0.0

        client_hours = sum(
            safe_float(r.get("hours", 0))
            for r in m_staff
            if project_map.get(r.get("project_id", ""), {}).get("case_type", "").strip().lower() == "client"
        )
        total_hours = sum(safe_float(r.get("hours", 0)) for r in m_staff)
        client_pct = (client_hours / total_hours * 100.0) if total_hours > 0 else 0.0
        other_pct = 100.0 - client_pct if total_hours > 0 else 0.0

        martha_rows.append({
            "month": month_name,
            "ftes": f"{avg_ftes:.1f}",
            "gross_cost": money_fmt(gross_cost_all),
            "recovery_billed_work": money_fmt(billed_recovery),
            "net_cost_to_practice": money_fmt(max(gross_cost_all - billed_recovery, 0.0)),
            "total_recovery_pct": pct_fmt(recovery_total_pct),
            "client_codes_pct": pct_fmt(client_pct),
            "other_practice_ip_pct": pct_fmt(other_pct)
        })

    key_wins = []
    key_challenges = []
    action_items = []

    if total_billing > 0:
        key_wins.append(f"Recovered {money_fmt(total_billing)} in billed work in {selected_period_label}.")
    if recovery_pct >= 100:
        key_wins.append("Recovery exceeded gross cost for the selected month.")

    if gross_cost > total_billing:
        key_challenges.append(f"Net cost to practice is {money_fmt(gross_cost - total_billing)} for {selected_period_label}.")
    if not month_billing:
        key_challenges.append("No billing entries found for the selected month.")

    action_items.append("Review teams with low billed recovery against staffed cost.")
    action_items.append("Validate cost rates for all employee designations to improve recovery accuracy.")
    action_items.append("Add live NPS and CTSU source if those KPIs are required on the dashboard.")

    # Case billing tracker (summarized by selected aggregation)
    def tracker_bucket_label(dt, project_id="", project_name_fallback=""):
        # For field-based grouping, determine the grouping field value
        if effective_tracker_summary in ["type_of_project", "team_classification_main", "team_classification_1", "team_classification_2", "region"]:
            project = project_map.get(project_id, {})
            team = team_map.get(project.get("team_id", ""), {})
            
            if effective_tracker_summary == "type_of_project":
                return project.get("project_type", "Unknown")
            elif effective_tracker_summary == "team_classification_main":
                return team.get("team_classification_main", "Unknown")
            elif effective_tracker_summary == "team_classification_1":
                return team.get("team_classification_1", "Unknown")
            elif effective_tracker_summary == "team_classification_2":
                return team.get("team_classification_2", "Unknown")
            elif effective_tracker_summary == "region":
                return project.get("office", "") or project.get("case_delivery_primary_location", "") or "Unknown"
        
        # Time-based grouping
        if effective_tracker_summary == "year":
            return str(dt.year)
        if effective_tracker_summary == "month":
            return dt.strftime("%b %Y")
        if effective_tracker_summary == "week":
            ws = dt - timedelta(days=dt.weekday())
            we = ws + timedelta(days=6)
            return f"{ws.strftime('%d %b %Y')} - {we.strftime('%d %b %Y')}"
        return selected_period_label

    tracker_grouped = {}
    for row in month_billing:
        dt = parse_date(row.get("date", ""))
        if not dt:
            continue
        project_id = row.get("project_id", "")
        project_name_fallback = row.get("project_name", "")
        bucket = tracker_bucket_label(dt, project_id, project_name_fallback)
        
        # For field-based grouping, use bucket as primary key; for project/time-based, use project+bucket
        if effective_tracker_summary in ["type_of_project", "team_classification_main", "team_classification_1", "team_classification_2", "region"]:
            tracker_key = (bucket,)
        else:
            tracker_key = (project_id or f"name::{project_name_fallback}", bucket)
        
        if tracker_key not in tracker_grouped:
            tracker_grouped[tracker_key] = {
                "project_id": project_id,
                "project_name_fallback": project_name_fallback,
                "case_code": row.get("case_code", ""),
                "billing_total": 0.0,
                "bucket": bucket,
                "project_ids": set(),
            }
        tracker_grouped[tracker_key]["billing_total"] += safe_float(row.get("billing_amount", 0))
        if not tracker_grouped[tracker_key]["case_code"]:
            tracker_grouped[tracker_key]["case_code"] = row.get("case_code", "")
        if project_id:
            tracker_grouped[tracker_key]["project_ids"].add(project_id)

    tracker_rows = []
    for _, grouped in tracker_grouped.items():
        project = project_map.get(grouped.get("project_id", ""), {})
        team = team_map.get(project.get("team_id", ""), {})
        manager_name = ""
        if team.get("manager_id"):
            manager = user_map.get(team.get("manager_id"), {})
            manager_name = f"{manager.get('first_name', '')} {manager.get('last_name', '')}".strip()

        project_dates = []
        for s in month_staffing:
            matches = False
            if grouped.get("project_id"):
                matches = s.get("project_id") == grouped.get("project_id")
            else:
                matches = (s.get("project_name", "") == grouped.get("project_name_fallback", ""))
            if matches:
                dt = parse_date(s.get("date", ""))
                if dt and tracker_bucket_label(dt, grouped.get("project_id", ""), grouped.get("project_name_fallback", "")) == grouped.get("bucket"):
                    project_dates.append(dt)

        duration_days = 0
        if project_dates:
            duration_days = (max(project_dates) - min(project_dates)).days + 1

        tracker_rows.append({
            "month": grouped.get("bucket", selected_period_label),
            "project_type": project.get("project_type", ""),
            "for_util": project.get("type_for_util", ""),
            "team_classification_main": team.get("team_classification_main", ""),
            "team_classification_1": team.get("team_classification_1", ""),
            "team_classification_2": team.get("team_classification_2", ""),
            "project_name": project.get("project_name", grouped.get("project_name_fallback", "")),
            "case_code": grouped.get("case_code", ""),
            "stakeholder": project.get("requestor", ""),
            "region": project.get("office", "") or project.get("case_delivery_primary_location", ""),
            "duration_days": duration_days,
            "billed_amount": money_fmt(grouped.get("billing_total", 0.0)),
            "bcn_manager": manager_name,
            "_billing_total_sort": grouped.get("billing_total", 0.0),
        })

    tracker_rows.sort(key=lambda r: r.get("_billing_total_sort", 0.0), reverse=True)
    for row in tracker_rows:
        row.pop("_billing_total_sort", None)

    metrics = {
        "scorecard": {
            "total_ftes": f"{total_ftes:.1f}",
            "fte_trend": fte_trend,
            "total_billing": money_fmt(total_billing),
            "billing_trend": billing_trend,
            "recovery": pct_fmt(recovery_pct),
            "recovery_target": "100%",
            "gender_ratio": gender_ratio,
            "female_share": female_share,
            "nps": nps_value,
            "nps_trend": nps_trend,
            "ctsu": ctsu_value,
            "ctsu_trend": ctsu_trend,
            "major_clients": major_clients,
            "notes": notes,
            "monthly_trend": monthly_trend
        },
        "billing_summary": {
            "global_sustainability": {
                "approved_headcount": approved_headcount_display(gs_team_ids),
                "approved_annual_budget": approved_annual_budget_display(gs_team_ids),
                "rows": global_sustainability_rows
            },
            "sr": {
                "approved_headcount": approved_headcount_display(sr_team_ids),
                "approved_annual_budget": approved_annual_budget_display(sr_team_ids),
                "rows": sr_rows
            }
        },
        "martha_summary": {
            "rows": martha_rows,
            "key_wins": key_wins,
            "key_challenges": key_challenges,
            "action_items": action_items
        },
        "case_billing_tracker": {
            "rows": tracker_rows
        }
    }

    return metrics

@app.route("/director")
@app.route("/director/home")
@login_required
@roles_required("DIRECTOR")
def director_home():
    allowed_views = {"scorecard", "billing", "martha", "tracker", "project_gantt", "chat"}
    view_candidates = request.args.getlist("view") or ["scorecard"]
    active_view = "scorecard"
    for candidate in reversed(view_candidates):
        cleaned = (candidate or "").strip().lower().split("?", 1)[0]
        if cleaned in allowed_views:
            active_view = cleaned
            break

    if "period_scope" in request.args:
        period_scope = request.args.get("period_scope", "monthly").strip().lower()
        session["director_period_scope"] = period_scope
    else:
        period_scope = session.get("director_period_scope", "monthly")

    if "tracker_summary" in request.args:
        tracker_summary = request.args.get("tracker_summary", "auto").strip().lower()
        session["director_tracker_summary"] = tracker_summary
    else:
        tracker_summary = session.get("director_tracker_summary", "auto")

    if tracker_summary not in ["auto", "project", "week", "month", "year", "type_of_project", "team_classification_main", "team_classification_1", "team_classification_2", "region"]:
        tracker_summary = "auto"

    if period_scope not in ["weekly", "monthly", "ytd"]:
        period_scope = "monthly"

    timeline_options = get_director_timeline_options(period_scope)
    timeline_values = {o["value"] for o in timeline_options}

    if "timeline" in request.args:
        selected_timeline = request.args.get("timeline", "").strip()
        session["director_timeline"] = selected_timeline
    elif "month" in request.args and period_scope == "monthly":
        selected_timeline = request.args.get("month", "").strip()
        session["director_timeline"] = selected_timeline
    else:
        selected_timeline = session.get("director_timeline", "")

    if (not selected_timeline) or (selected_timeline not in timeline_values):
        selected_timeline = timeline_options[-1]["value"] if timeline_options else ""
        session["director_timeline"] = selected_timeline
    
    metrics = None
    gantt_data = None
    manager_options = []
    team_class_main_options = []
    team_class_1_options = []
    team_class_2_options = []
    
    if active_view == "project_gantt":
        # Handle project gantt view
        view_type = "weekly" if period_scope == "weekly" else "monthly"
        
        anchor_str = request.args.get("anchor", "").strip()
        if period_scope == "weekly":
            try:
                default_anchor = datetime.strptime(selected_timeline, "%Y-%m-%d").date()
            except Exception:
                default_anchor = datetime.now().date()
        elif period_scope == "monthly":
            selected_year, selected_month_num = parse_selected_month(selected_timeline)
            default_anchor = datetime(selected_year, selected_month_num, 1).date()
        else:
            try:
                default_anchor = datetime(int(selected_timeline), 1, 1).date()
            except Exception:
                default_anchor = datetime.now().date().replace(day=1)
        try:
            anchor_date = datetime.strptime(anchor_str, "%Y-%m-%d").date() if anchor_str else default_anchor
        except ValueError:
            anchor_date = default_anchor
        
        # Get filter values
        manager_filter = request.args.get("manager", "All").strip()
        team_class_main_filter = request.args.get("team_class_main", "All").strip()
        team_class_1_filter = request.args.get("team_class_1", "All").strip()
        team_class_2_filter = request.args.get("team_class_2", "All").strip()
        
        gantt_data = build_director_gantt(
            active_view=view_type,
            anchor_date=anchor_date,
            manager_filter=manager_filter if manager_filter != "All" else None,
            team_class_main_filter=team_class_main_filter if team_class_main_filter != "All" else None,
            team_class_1_filter=team_class_1_filter if team_class_1_filter != "All" else None,
            team_class_2_filter=team_class_2_filter if team_class_2_filter != "All" else None
        )
        
        # Build filter options
        teams = get_teams()
        users = get_users()
        user_map = {u["id"]: u for u in users}
        
        # Managers
        manager_ids = {t.get("manager_id") for t in teams if t.get("manager_id")}
        manager_options = [{"value": "All", "label": "All"}]
        for mid in sorted(manager_ids):
            user = user_map.get(mid, {})
            full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or mid
            manager_options.append({"value": mid, "label": full_name})
        
        # Team classifications
        team_class_main_vals = {t.get("team_classification_main") for t in teams if t.get("team_classification_main")}
        team_class_main_options = ["All"] + sorted([v for v in team_class_main_vals if v])
        
        team_class_1_vals = {t.get("team_classification_1") for t in teams if t.get("team_classification_1")}
        team_class_1_options = ["All"] + sorted([v for v in team_class_1_vals if v])
        
        team_class_2_vals = {t.get("team_classification_2") for t in teams if t.get("team_classification_2")}
        team_class_2_options = ["All"] + sorted([v for v in team_class_2_vals if v])
    else:
        # Original director metrics views
        metrics = build_director_metrics(selected_timeline, period_scope, tracker_summary)
    
    return render_template(
        "director_home.html",
        metrics=metrics,
        gantt_data=gantt_data,
        selected_timeline=selected_timeline,
        timeline_options=timeline_options,
        active_view=active_view,
        manager_options=manager_options,
        team_class_main_options=team_class_main_options,
        team_class_1_options=team_class_1_options,
        team_class_2_options=team_class_2_options,
        manager_filter=request.args.get("manager", "All"),
        team_class_main_filter=request.args.get("team_class_main", "All"),
        team_class_1_filter=request.args.get("team_class_1", "All"),
        team_class_2_filter=request.args.get("team_class_2", "All"),
        view_type=("weekly" if period_scope == "weekly" else "monthly"),
        period_scope=period_scope,
        tracker_summary=tracker_summary,
        period_trend_label=("vs prior week" if period_scope == "weekly" else ("vs prior year" if period_scope == "ytd" else "vs prior month"))
    )


@app.route("/manager")
@app.route("/manager/home")
@login_required
@roles_required("MANAGER")
def manager_home():
    user = session["user"]
    manager_tab = request.args.get("tab", "dashboard").strip().lower()
    if manager_tab not in ["dashboard", "overview", "chat"]:
        manager_tab = "dashboard"

    teams = [t for t in get_teams() if t.get("manager_id") == user["id"]]
    billing_entries = get_billing_entries()
    staffing_entries = get_staffing_entries()
    employees = get_employees()
    projects = get_projects()

    employee_map = {e["id"]: e for e in employees}
    project_map = {p["id"]: p for p in projects}
    team_ids = {t["id"] for t in teams}

    today = datetime.now().date()
    current_year = today.year
    current_month = today.month

    manager_staffing = [s for s in staffing_entries if s.get("manager_id") == user["id"]]
    manager_billing = [b for b in billing_entries if b.get("manager_id") == user["id"]]

    # KPI calculations
    ytd_billing_value = 0.0
    for b in manager_billing:
        try:
            entry_date = datetime.strptime(b.get("date", ""), "%Y-%m-%d").date()
            if entry_date.year == current_year:
                ytd_billing_value += float(b.get("billing_amount", 0) or 0)
        except Exception:
            pass

    monthly_billing_value = 0.0
    for b in manager_billing:
        try:
            entry_date = datetime.strptime(b.get("date", ""), "%Y-%m-%d").date()
            if entry_date.year == current_year and entry_date.month == current_month:
                monthly_billing_value += float(b.get("billing_amount", 0) or 0)
        except Exception:
            pass

    monthly_staffing = []
    for s in manager_staffing:
        try:
            entry_date = datetime.strptime(s.get("date", ""), "%Y-%m-%d").date()
            if entry_date.year == current_year and entry_date.month == current_month:
                monthly_staffing.append(s)
        except Exception:
            pass

    working_days_with_entries = sorted({s.get("date") for s in monthly_staffing if s.get("date")})
    avg_ftes_value = 0.0
    if working_days_with_entries:
        daily_fte_totals = []
        for day in working_days_with_entries:
            day_hours = sum(float(s.get("hours", 0) or 0) for s in monthly_staffing if s.get("date") == day)
            daily_fte_totals.append(day_hours / 8.0)
        avg_ftes_value = sum(daily_fte_totals) / len(daily_fte_totals)

    manager_employee_ids = {
        e["id"] for e in employees
        if e.get("team_id") in team_ids
    }

    staffed_hours = sum(float(s.get("hours", 0) or 0) for s in monthly_staffing)

    month_start_for_util = today.replace(day=1)
    days_elapsed = [month_start_for_util + timedelta(days=i) for i in range((today - month_start_for_util).days + 1)]
    weekdays_elapsed = [d for d in days_elapsed if d.weekday() < 5]
    available_hours = len(manager_employee_ids) * len(weekdays_elapsed) * 8

    utilization_pct = 0.0
    if available_hours > 0:
        utilization_pct = (staffed_hours / available_hours) * 100

    def money_k(v):
        return f"${round(v/1000):.0f}K" if v >= 1000 else f"${v:,.0f}"

    kpis = [
        {"label": "Total YTD Manager Billing", "value": money_k(ytd_billing_value)},
        {"label": "Monthly Manager Billing", "value": money_k(monthly_billing_value)},
        {"label": "Avg FTEs", "value": f"{avg_ftes_value:.1f}"},
        {"label": "Utilization", "value": f"{utilization_pct:.0f}%"},
        {"label": "YTD NPS", "value": "-"},
        {"label": "YTD CTSU", "value": "-"},
    ]

    # Shared view controls
    active_view = request.args.get("view", "weekly").strip().lower()
    if active_view not in ["weekly", "monthly"]:
        active_view = "weekly"

    anchor_str = request.args.get("anchor", "").strip()
    try:
        anchor_date = datetime.strptime(anchor_str, "%Y-%m-%d").date() if anchor_str else today
    except ValueError:
        anchor_date = today

    if active_view == "weekly":
        period_start = anchor_date - timedelta(days=anchor_date.weekday())
        visible_dates = [period_start + timedelta(days=i) for i in range(5)]
        prev_anchor = period_start - timedelta(days=7)
        next_anchor = period_start + timedelta(days=7)
        period_label = f"{visible_dates[0].strftime('%d %b %Y')} - {visible_dates[-1].strftime('%d %b %Y')}"
    else:
        period_start = anchor_date.replace(day=1)
        if period_start.month == 12:
            next_month_start = period_start.replace(year=period_start.year + 1, month=1, day=1)
        else:
            next_month_start = period_start.replace(month=period_start.month + 1, day=1)

        period_end = next_month_start - timedelta(days=1)
        visible_dates = []
        cursor = period_start
        while cursor <= period_end:
            if cursor.weekday() < 5:
                visible_dates.append(cursor)
            cursor += timedelta(days=1)

        prev_anchor = (period_start - timedelta(days=1)).replace(day=1)
        next_anchor = next_month_start
        period_label = period_start.strftime("%B %Y")

    visible_date_strings = [d.strftime("%Y-%m-%d") for d in visible_dates]
    week_date_labels = [d.strftime("%d %b") for d in visible_dates]
    week_day_labels = [d.strftime("%a").upper() for d in visible_dates]

    manager_entries_in_range = [
        s for s in manager_staffing
        if s.get("date") in visible_date_strings
    ]

    overview_gantt = None
    overview_month_label = anchor_date.strftime("%b %Y")
    overview_working_days = working_days_in_month(anchor_date.year, anchor_date.month)
    overview_manager_tables = []

    if manager_tab == "overview":
        overview_gantt = build_director_gantt(active_view=active_view, anchor_date=anchor_date)

        month_start = anchor_date.replace(day=1)
        if month_start.month == 12:
            next_month_start = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            next_month_start = month_start.replace(month=month_start.month + 1, day=1)

        month_staffing_all = []
        for row in staffing_entries:
            dt = parse_date(row.get("date", ""))
            if dt and month_start <= dt < next_month_start:
                month_staffing_all.append(row)

        users = get_users()
        manager_users = sorted(
            [u for u in users if u.get("role") == "MANAGER"],
            key=lambda u: ((u.get("first_name", "") + " " + u.get("last_name", "")).strip().lower(), u.get("id", ""))
        )

        employee_day_stats = {}
        for row in month_staffing_all:
            dt = parse_date(row.get("date", ""))
            emp_id = row.get("employee_id", "")
            mgr_id = row.get("manager_id", "")
            if not dt or not emp_id:
                continue
            key = (emp_id, dt)
            stats = employee_day_stats.setdefault(key, {"hours": 0.0, "managers": set()})
            stats["hours"] += safe_float(row.get("hours", 0))
            if mgr_id:
                stats["managers"].add(mgr_id)

        overloaded_keys = {
            key
            for key, stats in employee_day_stats.items()
            if stats["hours"] > 8.0 and len(stats["managers"]) > 1
        }

        for mgr in manager_users:
            mgr_id = mgr.get("id", "")
            manager_name = f"{mgr.get('first_name', '')} {mgr.get('last_name', '')}".strip() or mgr_id
            manager_rows = [r for r in month_staffing_all if r.get("manager_id") == mgr_id]

            by_employee = {}
            for row in manager_rows:
                dt = parse_date(row.get("date", ""))
                emp_id = row.get("employee_id", "")
                if not dt or not emp_id:
                    continue
                emp = employee_map.get(emp_id, {})
                bucket = by_employee.setdefault(
                    emp_id,
                    {
                        "employee_name": emp.get("name", emp_id),
                        "total_hours": 0.0,
                        "dates": set(),
                        "overloaded_dates": set(),
                    },
                )
                bucket["total_hours"] += safe_float(row.get("hours", 0))
                bucket["dates"].add(dt)
                if (emp_id, dt) in overloaded_keys:
                    bucket["overloaded_dates"].add(dt)

            table_rows = []
            for emp_id, item in sorted(by_employee.items(), key=lambda x: x[1]["employee_name"].lower()):
                fte_days = round(item["total_hours"] / 8.0, 2)
                if abs(fte_days - float(overview_working_days)) < 1e-9:
                    coverage = f"{fte_days:g} days"
                else:
                    coverage = format_date_ranges(item["dates"])

                conflict_dates_text = format_date_ranges(item["overloaded_dates"]) if item["overloaded_dates"] else "-"

                table_rows.append(
                    {
                        "employee_id": emp_id,
                        "employee_name": item["employee_name"],
                        "total_hours": round(item["total_hours"], 2),
                        "fte_days": fte_days,
                        "coverage": coverage,
                        "has_overlap_conflict": bool(item["overloaded_dates"]),
                        "conflict_dates": conflict_dates_text,
                    }
                )

            overview_manager_tables.append(
                {
                    "manager_id": mgr_id,
                    "manager_name": manager_name,
                    "rows": table_rows,
                }
            )

    # Employees manager has ever worked with
    all_employee_ids_worked_with = []
    seen_emp_ids = set()
    for s in manager_staffing:
        emp_id = s.get("employee_id")
        if emp_id and emp_id not in seen_emp_ids:
            seen_emp_ids.add(emp_id)
            all_employee_ids_worked_with.append(emp_id)

    # Include employees from manager teams too
    for e in employees:
        if e.get("team_id") in team_ids and e["id"] not in seen_emp_ids:
            seen_emp_ids.add(e["id"])
            all_employee_ids_worked_with.append(e["id"])

    active_employee_ids = []
    inactive_employee_ids = []
    active_seen = set()

    for s in manager_entries_in_range:
        emp_id = s.get("employee_id")
        if emp_id and emp_id in seen_emp_ids and emp_id not in active_seen:
            active_seen.add(emp_id)
            active_employee_ids.append(emp_id)

    for emp_id in all_employee_ids_worked_with:
        if emp_id not in active_seen:
            inactive_employee_ids.append(emp_id)

    ordered_employee_ids = active_employee_ids + inactive_employee_ids

    # Staffing matrix table
    staffing_rows = []
    for emp_id in ordered_employee_ids:
        emp = employee_map.get(emp_id)
        if not emp:
            continue

        row_cells_project = []
        row_cells_staffing = []
        has_any_value = False

        for date_str in visible_date_strings:
            matching = [
                s for s in manager_entries_in_range
                if s.get("employee_id") == emp_id and s.get("date") == date_str
            ]

            if not matching:
                row_cells_project.append("")
                row_cells_staffing.append("")
                continue

            # Build project view
            cell_values_project = []
            cell_values_staffing = []
            
            for s in matching:
                hours = float(s.get("hours", 0) or 0)
                
                # Project view: show project name
                if s.get("project_id") and s.get("project_id") in project_map:
                    label = project_map[s["project_id"]].get("project_name", "")
                elif s.get("project_name"):
                    label = s.get("project_name", "")
                else:
                    label = s.get("staffing_type", "")
                
                display_project = f"{label} ({hours:g}h)" if label else f"{s.get('staffing_type', '')} ({hours:g}h)"
                if display_project not in cell_values_project:
                    cell_values_project.append(display_project)
                
                # Staffing type view: show staffing type
                staffing_type = s.get("staffing_type", "Unspecified")
                display_staffing = f"{staffing_type} ({hours:g}h)"
                if display_staffing not in cell_values_staffing:
                    cell_values_staffing.append(display_staffing)

            cell_text_project = ", ".join(cell_values_project)
            cell_text_staffing = ", ".join(cell_values_staffing)
            
            if cell_text_project or cell_text_staffing:
                has_any_value = True
            
            row_cells_project.append(cell_text_project)
            row_cells_staffing.append(cell_text_staffing)

        staffing_rows.append({
            "employee_id": emp_id,
            "name": emp.get("name", ""),
            "designation": emp.get("designation", ""),
            "has_activity": has_any_value,
            "week": row_cells_project,
            "week_staffing": row_cells_staffing
        })

    # Gantt with parallel projects per team (one row per team, multiple bars for overlapping projects)
    gantt_rows = []

    billing_in_range = [
        b for b in manager_billing
        if b.get("date") in visible_date_strings
    ]

    def resolve_project_meta(project_id, fallback_name=""):
        project = project_map.get(project_id, {}) if project_id else {}
        full_name = project.get("project_name", "") or fallback_name or "Unspecified"
        region = (
            project.get("billing_region", "")
            or project.get("office", "")
            or project.get("case_delivery_primary_location", "")
            or "Global"
        )
        return {
            "name": full_name,
            "region": region,
            "requestor": project.get("requestor", "") or "-",
        }

    for team in teams:
        team_entries = [
            s for s in manager_entries_in_range
            if resolve_staffing_team_id(s, project_map, employee_map) == team["id"]
        ]

        if not team_entries:
            gantt_rows.append({
                "team_name": team.get("team_name", ""),
                "bars": [],
                "is_empty": True
            })
            continue

        # Collect project data with day-level FTEs
        project_day_ftes = {}   # {project_key: {date_str: fte_on_day}}
        project_info = {}       # {project_key: {"name": str}}

        for s in team_entries:
            if s.get("project_id") and s.get("project_id") in project_map:
                project_key = s["project_id"]
                label = project_map[s["project_id"]].get("project_name", "")
            elif s.get("project_name"):
                project_key = s.get("project_name")
                label = s.get("project_name", "")
            else:
                project_key = s.get("staffing_type", "Unspecified")
                label = s.get("staffing_type", "Unspecified")

            if not label:
                continue

            date_key = s.get("date")
            hours = float(s.get("hours", 0) or 0)
            project_day_ftes.setdefault(project_key, {})
            project_day_ftes[project_key][date_key] = project_day_ftes[project_key].get(date_key, 0.0) + (hours / 8.0)
            if project_key not in project_info:
                meta = resolve_project_meta(s.get("project_id", ""), s.get("project_name", "") or label)
                project_info[project_key] = {
                    "name": meta["name"],
                    "region": meta["region"],
                    "requestor": meta["requestor"],
                    "total_hours": 0.0,
                    "total_billing": 0.0,
                }
            project_info[project_key]["total_hours"] += hours

        for b in billing_in_range:
            if b.get("project_id") and b.get("project_id") in project_map:
                project_key = b.get("project_id")
            elif b.get("project_name"):
                project_key = b.get("project_name")
            else:
                continue
            if project_key in project_info:
                project_info[project_key]["total_billing"] += float(b.get("billing_amount", 0) or 0)

        # Build occupied-day indexes for overlap detection
        project_indexes = {}
        for project_key, day_map in project_day_ftes.items():
            indexes_in_range = {
                visible_date_strings.index(d)
                for d, fte in day_map.items()
                if d in visible_date_strings and fte > 0
            }
            if indexes_in_range:
                project_indexes[project_key] = indexes_in_range

        # Detect overlapping projects and assign to levels (rows)
        def projects_overlap(idx_set_1, idx_set_2):
            return len(idx_set_1.intersection(idx_set_2)) > 0

        project_to_level = {}
        sorted_projects = sorted(project_indexes.items(), key=lambda x: min(x[1]))

        for project_key, idx_set in sorted_projects:
            level = 0
            for other_project_key, other_level in project_to_level.items():
                if other_level == level and projects_overlap(idx_set, project_indexes[other_project_key]):
                    level += 1
            project_to_level[project_key] = level

        # Build bars for each project
        bars = []
        for project_key in sorted(project_indexes.keys()):
            project_name = project_info[project_key]["name"]
            level = project_to_level[project_key]

            # Build segments (split when date is not contiguous OR when day FTE changes)
            day_fte_by_idx = {}
            for d, fte in project_day_ftes.get(project_key, {}).items():
                if d in visible_date_strings and fte > 0:
                    day_fte_by_idx[visible_date_strings.index(d)] = fte

            dates_for_project = sorted(day_fte_by_idx.keys())

            segments = []
            if dates_for_project:
                i = 0
                while i < len(dates_for_project):
                    start_idx = dates_for_project[i]
                    end_idx = start_idx
                    current_fte = day_fte_by_idx[start_idx]

                    # Find contiguous range with same day-level FTE
                    while (
                        i + 1 < len(dates_for_project)
                        and dates_for_project[i + 1] == dates_for_project[i] + 1
                        and abs(day_fte_by_idx[dates_for_project[i + 1]] - current_fte) < 1e-9
                    ):
                        i += 1
                        end_idx = dates_for_project[i]

                    span = end_idx - start_idx + 1
                    fte_label = f"{current_fte:.1f}".rstrip("0").rstrip(".")
                    label = f"{project_name} - {fte_label} FTEs"
                    segments.append({
                        "style": f"grid-column: {start_idx + 1} / span {span};",
                        "label": label
                    })
                    i += 1

            # Calculate daily FTE and total days from visible data
            num_days = len(day_fte_by_idx)
            if num_days > 0:
                # Average FTE across visible days
                daily_fte = round(sum(day_fte_by_idx.values()) / num_days, 2)
            else:
                daily_fte = 0.0

            bars.append({
                "project_key": project_key,
                "project_name": project_name,
                "level": level,
                "tooltip": {
                    "project_name": project_info[project_key].get("name", project_name),
                    "daily_ftes": daily_fte,
                    "total_days": num_days,
                    "region": project_info[project_key].get("region", "Global"),
                    "requestor": project_info[project_key].get("requestor", "-"),
                    "total_billing": f"${project_info[project_key].get('total_billing', 0.0):,.2f}",
                },
                "segments": segments
            })

        max_level = max([b["level"] for b in bars]) if bars else 0
        gantt_rows.append({
            "team_name": team.get("team_name", ""),
            "bars": bars,
            "is_empty": False,
            "max_level": max_level
        })

    return render_template(
        "manager_home.html",
        manager_tab=manager_tab,
        kpis=kpis,
        teams=teams,
        gantt_rows=gantt_rows,
        overview_gantt=overview_gantt,
        overview_month_label=overview_month_label,
        overview_working_days=overview_working_days,
        overview_manager_tables=overview_manager_tables,
        staffing_rows=staffing_rows,
        week_day_labels=week_day_labels,
        week_date_labels=week_date_labels,
        visible_dates=visible_date_strings,
        period_label=period_label,
        active_view=active_view,
        prev_anchor=prev_anchor.strftime("%Y-%m-%d"),
        next_anchor=next_anchor.strftime("%Y-%m-%d"),
        current_anchor=anchor_date.strftime("%Y-%m-%d"),
        greeting=get_greeting(),
        employees=employees,
        projects=projects,
        reports=get_reports(),
        users=get_users(),
        all_teams=get_teams()
    )

@app.route("/manager/projects")
@login_required
@roles_required("MANAGER")
def manager_projects():
    user = session["user"]
    rows = [p for p in get_projects() if p.get("created_by") == user["id"]]
    return render_template("manager_projects.html", rows=rows, teams=get_teams(), users=get_users())

@app.route("/manager/projects/edit/<item_id>", methods=["GET", "POST"])
@login_required
@roles_required("MANAGER")
def manager_projects_edit(item_id):
    rows = get_projects()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Project not found.", "error")
        return redirect(url_for("manager_projects"))

    if not can_edit_project(session.get("user"), item):
        flash("You do not have access to edit this project.", "error")
        return redirect(url_for("manager_projects"))

    if request.method == "POST":
        for key in FULL_PROJECT_FIELDS:
            item[key] = request.form.get(key, "").strip()

        save_projects(rows)
        flash("Project updated.", "success")
        return redirect(url_for("manager_projects"))

    return render_template(
        "manager_project_edit.html",
        item=item,
        teams=get_teams(),
        users=get_users()
    )

@app.route("/manager/projects/delete/<item_id>")
@login_required
@roles_required("MANAGER")
def manager_projects_delete(item_id):
    rows = get_projects()
    item = find_by_id(rows, item_id)

    if not item:
        flash("Project not found.", "error")
        return redirect(url_for("manager_projects"))

    if not can_edit_project(session.get("user"), item):
        flash("You do not have access to delete this project.", "error")
        return redirect(url_for("manager_projects"))

    save_projects([r for r in rows if r.get("id") != item_id])
    flash("Project deleted.", "success")
    return redirect(url_for("manager_projects"))

@app.route("/api/staffing/prefill")
@login_required
@roles_required("MANAGER")
def api_staffing_prefill():
    manager_id = session["user"]["id"]

    requested_date = request.args.get("date", "").strip()
    load_date = request.args.get("load_date", "").strip()

    # Opening modal should use current date by default.
    # Custom "Load From Date" can still override via load_date.
    target_date = load_date or requested_date or datetime.now().strftime("%Y-%m-%d")

    staffing = get_staffing_entries()
    billing = get_billing_entries()
    employees = {e["id"]: e for e in get_employees()}
    projects = {p["id"]: p for p in get_projects()}

    staffing_rows = []
    for item in staffing:
        if item.get("manager_id") == manager_id and item.get("date") == target_date:
            project = projects.get(item.get("project_id", ""), {})
            staffing_rows.append({
                "employee_id": item.get("employee_id", ""),
                "employee_name": employees.get(item.get("employee_id", ""), {}).get("name", ""),
                "staffing_type": item.get("staffing_type", ""),
                "project_id": item.get("project_id", ""),
                "team_id": item.get("team_id", ""),
                "project_name": project.get("project_name", item.get("project_name", "")),
                "case_code": item.get("case_code", ""),
                "hours": float(item.get("hours", 0) or 0),
                "comments": item.get("comments", "")
            })

    # IMPORTANT:
    # Return saved billing rows from database for that same date,
    # not re-derived rows, so user sees latest saved edits.
    billing_rows = []
    for item in billing:
        if item.get("manager_id") == manager_id and item.get("date") == target_date:
            billing_rows.append({
                "project_id": item.get("project_id", ""),
                "project_name": item.get("project_name", ""),
                "project_type": item.get("project_type", ""),
                "case_code": item.get("case_code", ""),
                "billable_ftes": float(item.get("billable_ftes", 0) or 0),
                "billing_amount": float(item.get("billing_amount", 0) or 0),
                "comments": item.get("comments", "")
            })

    # If staffing exists but billing does not yet exist for that date,
    # derive once as fallback.
    if staffing_rows and not billing_rows:
        billing_rows = derive_billing_rows(staffing_rows)

    return jsonify({
        "source_date": target_date,
        "rows": staffing_rows,
        "billing_rows": billing_rows
    })

@app.route("/api/staffing/save", methods=["POST"])
@login_required
@roles_required("MANAGER")
def api_staffing_save():
    payload = request.get_json(force=True)
    date = (payload.get("date") or "").strip()
    rows = payload.get("rows", [])
    billing_rows = payload.get("billing_rows", [])
    manager_id = session["user"]["id"]

    if not date:
        return jsonify({"ok": False, "message": "Date is required."}), 400

    # Replace staffing entries for this manager/date
    staffing = [
        r for r in get_staffing_entries()
        if not (r.get("manager_id") == manager_id and r.get("date") == date)
    ]

    projects = {p["id"]: p for p in get_projects()}

    for row in rows:
        project = projects.get(row.get("project_id", ""), {})
        project_team_id = project.get("team_id", "")
        staffing.append({
            "id": next_id("SE", staffing),
            "date": date,
            "manager_id": manager_id,
            "employee_id": row.get("employee_id", ""),
            "staffing_type": row.get("staffing_type", ""),
            "project_id": row.get("project_id", ""),
            "project_name": row.get("project_name", ""),
            "team_id": project_team_id,
            "case_code": row.get("case_code", ""),
            "hours": float(row.get("hours", 0) or 0),
            "comments": row.get("comments", "")
        })

    save_staffing_entries(staffing)

    # Replace billing entries for this manager/date
    billing = [
        r for r in get_billing_entries()
        if not (r.get("manager_id") == manager_id and r.get("date") == date)
    ]

    for row in billing_rows:
        billing.append({
            "id": next_id("BE", billing),
            "date": date,
            "manager_id": manager_id,
            "project_id": row.get("project_id", ""),
            "project_name": row.get("project_name", ""),
            "project_type": row.get("project_type", ""),
            "case_code": row.get("case_code", ""),
            "billable_ftes": float(row.get("billable_ftes", 0) or 0),
            "billing_amount": float(row.get("billing_amount", 0) or 0),
            "comments": row.get("comments", "")
        })

    save_billing_entries(billing)

    return jsonify({"ok": True, "message": "Staffing and billing saved."})

@app.route("/api/staffing/save-multiple", methods=["POST"])
@login_required
@roles_required("MANAGER")
def api_staffing_save_multiple():
    """
    Save staffing and billing entries to multiple dates.
    Mode can be 'replace' (remove existing entries) or 'add' (append to existing).
    """
    payload = request.get_json(force=True)
    dates = payload.get("dates", [])
    rows = payload.get("rows", [])
    billing_rows = payload.get("billing_rows", [])
    mode = payload.get("mode", "replace").strip().lower()
    manager_id = session["user"]["id"]

    if not dates:
        return jsonify({"ok": False, "message": "No dates selected."}), 400

    if not rows:
        return jsonify({"ok": False, "message": "No staffing rows to save."}), 400

    projects = {p["id"]: p for p in get_projects()}
    staffing_entries = get_staffing_entries()
    billing_entries = get_billing_entries()

    for date in dates:
        date = date.strip()
        if not date:
            continue

        # Handle staffing entries
        if mode == "replace":
            # Remove existing entries for this manager/date
            staffing_entries = [
                r for r in staffing_entries
                if not (r.get("manager_id") == manager_id and r.get("date") == date)
            ]
        
        # Add new staffing entries for this date
        for row in rows:
            project = projects.get(row.get("project_id", ""), {})
            project_team_id = project.get("team_id", "")
            staffing_entries.append({
                "id": next_id("SE", staffing_entries),
                "date": date,
                "manager_id": manager_id,
                "employee_id": row.get("employee_id", ""),
                "staffing_type": row.get("staffing_type", ""),
                "project_id": row.get("project_id", ""),
                "project_name": row.get("project_name", ""),
                "team_id": project_team_id,
                "case_code": row.get("case_code", ""),
                "hours": float(row.get("hours", 0) or 0),
                "comments": row.get("comments", "")
            })

        # Handle billing entries
        if mode == "replace":
            # Remove existing billing for this manager/date
            billing_entries = [
                r for r in billing_entries
                if not (r.get("manager_id") == manager_id and r.get("date") == date)
            ]

        # Add new billing entries for this date
        for row in billing_rows:
            billing_entries.append({
                "id": next_id("BE", billing_entries),
                "date": date,
                "manager_id": manager_id,
                "project_id": row.get("project_id", ""),
                "project_name": row.get("project_name", ""),
                "project_type": row.get("project_type", ""),
                "case_code": row.get("case_code", ""),
                "billable_ftes": float(row.get("billable_ftes", 0) or 0),
                "billing_amount": float(row.get("billing_amount", 0) or 0),
                "comments": row.get("comments", "")
            })

    # Save all changes
    save_staffing_entries(staffing_entries)
    save_billing_entries(billing_entries)

    action = "replaced" if mode == "replace" else "added to"
    return jsonify({
        "ok": True,
        "message": f"Staffing and billing {action} {len(dates)} date(s) successfully."
    })

@app.route("/api/staffing/derive-billing", methods=["POST"])
@login_required
@roles_required("MANAGER")
def api_staffing_derive_billing():
    payload = request.get_json(force=True); return jsonify({"billing_rows": derive_billing_rows(payload.get("rows", []))})

@app.route("/api/projects/create", methods=["POST"])
@login_required
@roles_required("MANAGER", "ADMIN")
def api_projects_create():
    rows = get_projects()
    item = {"id": next_id("P", rows)}

    for key in FULL_PROJECT_FIELDS:
        item[key] = request.form.get(key, "").strip()

    item["created_by"] = session["user"]["id"]

    rows.append(item)
    save_projects(rows)
    return jsonify({"ok": True, "project": item})

@app.route("/api/reports/insync/generate", methods=["POST"])
@login_required
def api_insync_generate():
    """Generate an Insync report for a given month."""
    payload = request.get_json(force=True)
    month = payload.get("month", "")
    
    if not month:
        return jsonify({"ok": False, "message": "Month is required"})
    
    # Check if report already exists for this month
    reports = get_reports()
    existing_report = next((r for r in reports if r.get("type") == "monthly_insync" and r.get("month") == month), None)
    
    try:
        # Validate that report can be generated for this month.
        generate_insync_report_bytes(month)

        if existing_report:
            report_id = existing_report["id"]
            existing_report["generated_on"] = datetime.now().isoformat()
            existing_report["file_name"] = f"insync_{month}.xlsx"
            if "blob_pathname" in existing_report:
                del existing_report["blob_pathname"]
        else:
            report_id = next_id("RPT", reports)
            reports.append({
                "id": report_id,
                "type": "monthly_insync",
                "month": month,
                "generated_on": datetime.now().isoformat(),
                "file_name": f"insync_{month}.xlsx"
            })
        
        save_reports(reports)
        return jsonify({"ok": True, "message": "Report generated successfully", "report_id": report_id})
    
    except Exception as e:
        return jsonify({"ok": False, "message": f"Error generating report: {str(e)}"})

@app.route("/api/reports/insync/list")
@login_required
def api_insync_list():
    """Get all Insync reports sorted reverse chronologically by month."""
    reports = get_reports()
    insync_reports = [r for r in reports if r.get("type") == "monthly_insync"]
    
    # Sort reverse chronologically by month (so newer months first)
    insync_reports.sort(key=lambda x: x.get("month", ""), reverse=True)
    
    return jsonify({"ok": True, "reports": insync_reports})

@app.route("/api/reports/insync/download/<item_id>")
@login_required
def api_insync_download(item_id):
    """Download an Insync report file."""
    reports = get_reports()
    report = next((r for r in reports if r.get("id") == item_id and r.get("type") == "monthly_insync"), None)
    
    if not report:
        flash("Report not found.", "error")
        return redirect(url_for("route_by_role"))
    
    try:
        file_content = generate_insync_report_bytes(report.get("month", ""))
        return send_file(
            BytesIO(file_content),
            as_attachment=True,
            download_name=report.get("file_name") or "report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        flash(f"Unable to generate report download: {str(e)}", "error")
        return redirect(request.referrer or url_for("route_by_role"))

def generate_insync_report_bytes(month):
    """
    Generate an Insync report Excel file for the given month based on Insync.xlsx template.
    
    Args:
        month: Month string in format "YYYY-MM"
    
    Returns:
        The stored filename relative to UPLOAD_DIR
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for generating Insync reports. Please install it.")

    try:
        month_start = datetime.strptime(month, "%Y-%m")
    except ValueError:
        raise ValueError("Month must be in YYYY-MM format")

    user = session.get("user")
    if not user:
        raise ValueError("No user session found")

    manager_id = user.get("id", "")
    report_year = month_start.year
    report_month = month_start.month
    manager_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()

    teams = [t for t in get_teams() if t.get("manager_id") == manager_id]
    team_ids = {t.get("id", "") for t in teams}

    all_users = get_users()
    all_employees = get_employees()
    all_projects = get_projects()
    staffing_entries = get_staffing_entries()
    billing_entries = get_billing_entries()
    billing_rates = get_billing_rates()

    user_map = {u.get("id", ""): u for u in all_users}
    employee_map = {e.get("id", ""): e for e in all_employees}
    project_map = {p.get("id", ""): p for p in all_projects}
    project_by_name = {p.get("project_name", "").strip().lower(): p for p in all_projects if p.get("project_name")}

    def resolve_project(row):
        pid = row.get("project_id", "")
        if pid and pid in project_map:
            return project_map[pid]
        pname = row.get("project_name", "").strip().lower()
        return project_by_name.get(pname)

    def resolve_team_id_from_row(row):
        return resolve_staffing_team_id(row, project_map=project_map, employee_map=employee_map)

    rate_map = {}
    for rate in billing_rates:
        rate_map[(rate.get("region", "Global"), rate.get("project_type", ""))] = safe_float(rate.get("per_fte_rate", 0))

    def get_project_rate(project):
        if not project:
            return 0.0
        project_type = project.get("project_type", "")
        region = project.get("region") or project.get("office") or "Global"
        return rate_map.get((region, project_type), rate_map.get(("Global", project_type), 0.0))

    manager_month_staffing = [
        s for s in staffing_entries
        if s.get("manager_id") == manager_id and s.get("date", "").startswith(month)
    ]
    manager_month_billing = [
        b for b in billing_entries
        if b.get("manager_id") == manager_id and b.get("date", "").startswith(month)
    ]

    # Maintain a stable employee-column order for this manager for the selected year.
    manager_year_staffing = []
    for row in staffing_entries:
        if row.get("manager_id") != manager_id:
            continue
        dt = parse_date(row.get("date", ""))
        if dt and dt.year == report_year:
            manager_year_staffing.append(row)

    manager_year_staffing.sort(key=lambda r: (r.get("date", ""), r.get("id", "")))
    current_employee_order = []
    seen_emp_ids = set()
    for row in manager_year_staffing:
        emp_id = row.get("employee_id", "")
        if emp_id and emp_id not in seen_emp_ids:
            seen_emp_ids.add(emp_id)
            current_employee_order.append(emp_id)

    insync_orders = get_insync_employee_orders()
    order_key = f"{manager_id}:{report_year}"
    saved_order = [emp_id for emp_id in insync_orders.get(order_key, []) if emp_id]

    merged_order = []
    seen_merged = set()
    for emp_id in saved_order + current_employee_order:
        if emp_id and emp_id not in seen_merged:
            seen_merged.add(emp_id)
            merged_order.append(emp_id)

    insync_orders[order_key] = merged_order
    save_insync_employee_orders(insync_orders)

    employee_headers = [employee_map.get(emp_id, {}).get("name", emp_id) for emp_id in merged_order]

    def normalized_project_key(project_id, project_name):
        return project_id or f"name::{(project_name or '').strip().lower()}"

    billing_amount_by_date_project = {}
    for row in manager_month_billing:
        key = (row.get("date", ""), normalized_project_key(row.get("project_id", ""), row.get("project_name", "")))
        billing_amount_by_date_project[key] = billing_amount_by_date_project.get(key, 0.0) + safe_float(row.get("billing_amount", 0))

    wb = openpyxl.Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    header_fill = PatternFill(start_color="1f2328", end_color="1f2328", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    def write_headers(ws, headers):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx <= 10:
                ws.column_dimensions[cell.column_letter].width = 18
            else:
                ws.column_dimensions[cell.column_letter].width = 22
        ws.freeze_panes = "A2"

    def make_sheet_name(base_name, used_names):
        clean = (base_name or "Team").replace("/", "-").replace("\\", "-").replace("?", "-").replace("*", "-").replace("[", "(").replace("]", ")")
        clean = clean[:31] if clean else "Team"
        candidate = clean
        suffix = 2
        while candidate in used_names:
            tail = f"_{suffix}"
            candidate = f"{clean[:31-len(tail)]}{tail}"
            suffix += 1
        used_names.add(candidate)
        return candidate

    used_sheet_names = set()

    # Projects tab
    project_tab_headers = [
        "Project Name (Dropdown)",
        "Billing Case code",
        "Client Case Code",
        "Work Description (Be specific and explanatory about the work tasks)",
        "Product (Dropdown)",
        "Requestor",
        "NPS Contact",
        "Case Status (Dropdown)",
        "NPS Status (Dropdown)",
        "Case Manager against whom NPS will be reported",
        "BCN Case Execution Location",
        "Billed to end client as Fees/ Expense",
        "Case type",
        "Office",
        "Case Manager/ Principal",
        "Client name",
        "Case Partner",
        "Industry (Dropdown)",
        "Capability (Dropdown)",
    ]

    ws_projects = wb.create_sheet(title=make_sheet_name("Projects", used_sheet_names))
    write_headers(ws_projects, project_tab_headers)

    project_keys_in_month = set()
    for row in manager_month_staffing + manager_month_billing:
        key = normalized_project_key(row.get("project_id", ""), row.get("project_name", ""))
        project_keys_in_month.add(key)

    month_projects = []
    for key in project_keys_in_month:
        project = None
        if key.startswith("name::"):
            name_key = key.replace("name::", "", 1)
            project = project_by_name.get(name_key)
        else:
            project = project_map.get(key)
        if project:
            month_projects.append(project)

    month_projects.sort(key=lambda p: p.get("project_name", "").lower())

    project_row_idx = 2
    for project in month_projects:
        ws_projects.cell(row=project_row_idx, column=1).value = project.get("project_name", "")
        ws_projects.cell(row=project_row_idx, column=2).value = project.get("billing_case_code", "")
        ws_projects.cell(row=project_row_idx, column=3).value = project.get("client_case_code", "")
        ws_projects.cell(row=project_row_idx, column=4).value = project.get("work_description", "")
        ws_projects.cell(row=project_row_idx, column=5).value = project.get("product", "")
        ws_projects.cell(row=project_row_idx, column=6).value = project.get("requestor", "")
        ws_projects.cell(row=project_row_idx, column=7).value = project.get("nps_contact", "")
        ws_projects.cell(row=project_row_idx, column=8).value = project.get("case_status", "")
        ws_projects.cell(row=project_row_idx, column=9).value = project.get("nps_status", "")
        ws_projects.cell(row=project_row_idx, column=10).value = project.get("case_manager_for_nps", "")
        ws_projects.cell(row=project_row_idx, column=11).value = project.get("bcn_case_execution_location", "")
        ws_projects.cell(row=project_row_idx, column=12).value = project.get("billed_to_end_client", "")
        ws_projects.cell(row=project_row_idx, column=13).value = project.get("case_type", "")
        ws_projects.cell(row=project_row_idx, column=14).value = project.get("office", "")
        ws_projects.cell(row=project_row_idx, column=15).value = project.get("case_manager_principal", "")
        ws_projects.cell(row=project_row_idx, column=16).value = project.get("client_name", "")
        ws_projects.cell(row=project_row_idx, column=17).value = project.get("case_partner", "")
        ws_projects.cell(row=project_row_idx, column=18).value = project.get("industry", "")
        ws_projects.cell(row=project_row_idx, column=19).value = project.get("capability", "")
        project_row_idx += 1

    if project_row_idx == 2:
        ws_projects.cell(row=2, column=1).value = "No projects found for this month"

    # Team tabs
    team_headers = [
        "Cluster",
        "CoE",
        "BU (Dropdown)",
        "Team Code (Dropdown)",
        "CoE Lead + Director (Dropdown)",
        "S/TM Manager (Dropdown)",
        "Team Name for CTSU",
        "CTSU Ombudsperson",
        "Year",
        "Month",
        "Day",
        "Type",
        "Date",
        "Project Name (Dropdown)",
        "Billing Case code",
        "Client Case Code",
        "Billed Team Size",
        "Actual Billing",
        "Potential Billing",
    ] + employee_headers

    teams_sorted = sorted(teams, key=lambda t: t.get("team_name", "").lower())

    for team in teams_sorted:
        ws_team = wb.create_sheet(title=make_sheet_name(team.get("team_name", "Team"), used_sheet_names))
        write_headers(ws_team, team_headers)

        team_staffing_rows = [
            row for row in manager_month_staffing
            if resolve_team_id_from_row(row) == team.get("id", "")
        ]

        grouped = {}
        for row in team_staffing_rows:
            date = row.get("date", "")
            project_id = row.get("project_id", "")
            project_name = row.get("project_name", "")
            key = (date, normalized_project_key(project_id, project_name))
            bucket = grouped.setdefault(key, {"date": date, "project_id": project_id, "project_name": project_name, "rows": []})
            bucket["rows"].append(row)

        group_rows = list(grouped.values())
        group_rows.sort(key=lambda g: (g.get("date", ""), g.get("project_name", "").lower()))

        team_row_idx = 2
        for group in group_rows:
            date = group.get("date", "")
            project = resolve_project(group) or {}

            date_obj = parse_date(date)
            year_val = date_obj.year if date_obj else report_year
            month_val = date_obj.month if date_obj else report_month
            day_val = date_obj.day if date_obj else ""
            day_type = "Workday" if (date_obj and date_obj.weekday() < 5) else ("Weekend" if date_obj else "")

            ombudsperson_name = ""
            ombud_emp_id = team.get("ombudsperson_employee_id", "")
            if ombud_emp_id in employee_map:
                ombudsperson_name = employee_map[ombud_emp_id].get("name", "")

            coe_lead_director_name = ""
            director_id = team.get("director_id", "")
            if director_id and director_id in user_map:
                director_user = user_map[director_id]
                coe_lead_director_name = f"{director_user.get('first_name', '')} {director_user.get('last_name', '')}".strip()

            team_size = len({r.get("employee_id", "") for r in group["rows"] if r.get("employee_id", "")})
            total_hours = sum(safe_float(r.get("hours", 0)) for r in group["rows"])

            billing_key = (date, normalized_project_key(group.get("project_id", ""), group.get("project_name", "")))
            actual_billing = round(billing_amount_by_date_project.get(billing_key, 0.0), 2)
            potential_billing = round((total_hours / 8.0) * get_project_rate(project), 2)

            ws_team.cell(row=team_row_idx, column=1).value = team.get("team_classification_main", "")
            ws_team.cell(row=team_row_idx, column=2).value = team.get("team_classification_1", "")
            ws_team.cell(row=team_row_idx, column=3).value = team.get("team_classification_2", "")
            ws_team.cell(row=team_row_idx, column=4).value = team.get("id", "")
            ws_team.cell(row=team_row_idx, column=5).value = coe_lead_director_name
            ws_team.cell(row=team_row_idx, column=6).value = manager_name
            ws_team.cell(row=team_row_idx, column=7).value = team.get("team_name", "")
            ws_team.cell(row=team_row_idx, column=8).value = ombudsperson_name
            ws_team.cell(row=team_row_idx, column=9).value = year_val
            ws_team.cell(row=team_row_idx, column=10).value = month_val
            ws_team.cell(row=team_row_idx, column=11).value = day_val
            ws_team.cell(row=team_row_idx, column=12).value = day_type
            ws_team.cell(row=team_row_idx, column=13).value = date
            ws_team.cell(row=team_row_idx, column=14).value = project.get("project_name", group.get("project_name", ""))
            ws_team.cell(row=team_row_idx, column=15).value = project.get("billing_case_code", "")
            ws_team.cell(row=team_row_idx, column=16).value = project.get("client_case_code", "")
            ws_team.cell(row=team_row_idx, column=17).value = team_size
            ws_team.cell(row=team_row_idx, column=18).value = actual_billing
            ws_team.cell(row=team_row_idx, column=19).value = potential_billing

            row_by_employee = {}
            for row in group["rows"]:
                emp_id = row.get("employee_id", "")
                if not emp_id:
                    continue
                row_by_employee.setdefault(emp_id, []).append(row)

            for emp_idx, emp_id in enumerate(merged_order):
                cell_col = 20 + emp_idx
                emp_rows = row_by_employee.get(emp_id, [])
                if not emp_rows:
                    ws_team.cell(row=team_row_idx, column=cell_col).value = ""
                    continue

                staffing_types = []
                for emp_row in emp_rows:
                    st = (emp_row.get("staffing_type", "") or "").strip()
                    if st and st not in staffing_types:
                        staffing_types.append(st)

                if "Regular Hours" in staffing_types:
                    ws_team.cell(row=team_row_idx, column=cell_col).value = "Regular Hours"
                elif staffing_types:
                    ws_team.cell(row=team_row_idx, column=cell_col).value = ", ".join(staffing_types)
                else:
                    ws_team.cell(row=team_row_idx, column=cell_col).value = ""

            team_row_idx += 1

        if team_row_idx == 2:
            ws_team.cell(row=2, column=1).value = "No staffing data for this team in selected month"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

@app.route("/api/chat", methods=["POST"])
@login_required
def api_chat():
    """
    Multi-agent chat system:
    1. Query Understanding Agent - analyzes question and creates filter/cut strategy
    2. Code Generation Agent - writes Python code to process data
    3. Analysis Agent - uses processed data to answer (table or insights)
    4. Formatting Agent - formats the final answer
    """
    try:
        data = request.get_json()
        user_query = data.get("message", "").strip()
        
        if not user_query:
            return jsonify({"error": "No message provided"}), 400
        
        # Get OpenAI API key from environment variable
        openai_api_key = os.environ.get("OPENAI_API_KEY")
        if not openai_api_key:
            return jsonify({"error": "OpenAI API key not configured. Please set OPENAI_API_KEY environment variable."}), 500
        
        client = openai.OpenAI(api_key=openai_api_key)
        
        # Track progress for user feedback
        progress_steps = []
        
        # Gather all relevant data for context
        progress_steps.append({"agent": "System", "status": "Reading data..."})
        users = get_users()
        teams = get_teams()
        employees = get_employees()
        projects = get_projects()
        billing_entries = get_billing_entries()
        staffing_entries = get_staffing_entries()
        billing_rates = get_billing_rates()
        cost_rates = get_cost_rates()
        
        # Get current user info
        current_user = session.get("user", {})
        user_role = current_user.get("role", "")
        
        # Filter data based on user role
        if user_role == "MANAGER":
            manager_teams = [t for t in teams if t.get("manager_id") == current_user.get("id")]
            team_ids = {t.get("id") for t in manager_teams}
            projects = [p for p in projects if p.get("team_id") in team_ids]
            project_ids = {p.get("id") for p in projects}
            billing_entries = [b for b in billing_entries if b.get("project_id") in project_ids]
            staffing_entries = [s for s in staffing_entries if s.get("project_id") in project_ids]
        
        # Data schema description
        data_schema = {
            "users": "List[Dict] with keys: id, username, first_name, last_name, email, role (ADMIN/DIRECTOR/MANAGER)",
            "teams": "List[Dict] with keys: id, team_name, manager_id, description",
            "employees": "List[Dict] with keys: id, name, email, designation, team_id, status",
            "projects": "List[Dict] with keys: id, project_name, case_code, team_id, manager_id, requestor, billing_region, type_of_project, team_classification_main, team_classification_1, team_classification_2, status",
            "billing_entries": "List[Dict] with keys: id, project_id, project_name, case_code, date (YYYY-MM-DD string), billing_amount (float), billable_ftes, project_type, manager_id, notes. IMPORTANT: Each entry represents ONE DAY of billable work on a project. Use case_code field to filter by case code (e.g., Z5LB, J2RC). When counting entries, report as 'days worked' or 'billing days', NOT 'cases'.",
            "staffing_entries": "List[Dict] with keys: id, employee_id, project_id, project_name, date (YYYY-MM-DD string), hours (float), staffing_type (e.g., Regular Hour, Sick Leave). Each entry represents one day of staffing.",
            "billing_rates": "List[Dict] with keys: id, employee_id, designation, rate_per_day",
            "cost_rates": "List[Dict] with keys: id, employee_id, designation, cost_per_day"
        }
        
        # ===== AGENT 1: Query Understanding =====
        progress_steps.append({"agent": "Agent 1", "status": "Understanding query..."})
        agent1_prompt = f"""You are Agent 1: Query Understanding Agent.
Analyze the user's question and determine:
1. What data sources are needed (users, teams, employees, projects, billing_entries, staffing_entries, billing_rates, cost_rates)
2. What filters should be applied (e.g., date ranges, specific teams, project types)
3. What aggregations/groupings are needed (e.g., sum by month, group by manager)
4. Whether the user wants a table output or insights/analysis

Available data schema:
{json.dumps(data_schema, indent=2)}

User Question: "{user_query}"

Respond in JSON format:
{{
  "data_sources": ["list", "of", "sources"],
  "filters": {{"description": "what filters to apply"}},
  "aggregations": {{"description": "how to group/aggregate"}},
  "output_type": "table" or "insights"
}}"""

        response1 = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": agent1_prompt}],
            response_format={"type": "json_object"},
            max_tokens=800,
            temperature=0.3
        )
        
        query_plan = json.loads(response1.choices[0].message.content)
        
        # ===== AGENT 2: Code Generation =====
        progress_steps.append({"agent": "Agent 2", "status": "Filtering data..."})
        agent2_prompt = f"""You are Agent 2: Code Generation Agent.
Based on the query plan, write Python code to process the data.

Query Plan:
{json.dumps(query_plan, indent=2)}

User Question: "{user_query}"

Available variables in scope:
- users, teams, employees, projects, billing_entries, staffing_entries, billing_rates, cost_rates (all are List[Dict])
- datetime module is available via code: from datetime import datetime

Data Access Examples:
- Filter by case_code: [b for b in billing_entries if b.get('case_code') == 'Z5LB']
- Filter by month: [b for b in billing_entries if b.get('date','').startswith('2026-03')]
- Sum billing amounts: sum(float(b.get('billing_amount', 0)) for b in filtered_data)

Requirements:
1. Write clean Python code that processes the data according to the plan
2. Store the final result in a variable called 'result'
3. 'result' should be a list of dictionaries (table rows) or a dictionary with summary data
4. Use pandas if needed: import pandas as pd
5. Handle date parsing carefully: dates are strings in format 'YYYY-MM-DD'
6. Include helpful column names in the result
7. When filtering by case code, use b.get('case_code') field from billing_entries
8. For monthly filters like "March", filter dates that start with the year-month (e.g., '2026-03')

Example output format for tables:
result = [
  {{"month": "Jan", "total": 1000, "count": 5}},
  {{"month": "Feb", "total": 1200, "count": 6}}
]

Write ONLY the Python code, no explanations:"""

        response2 = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": agent2_prompt}],
            max_tokens=1500,
            temperature=0.2
        )
        
        generated_code = response2.choices[0].message.content.strip()
        # Remove code fences if present
        if generated_code.startswith("```python"):
            generated_code = generated_code.split("```python")[1].split("```")[0].strip()
        elif generated_code.startswith("```"):
            generated_code = generated_code.split("```")[1].split("```")[0].strip()
        
        # Execute the generated code safely
        progress_steps.append({"agent": "Agent 2", "status": "Processing data..."})
        
        # Import modules before exec
        import datetime
        import pandas as pd
        
        safe_globals = {
            "__builtins__": {
                "__import__": __import__,
                "len": len, "sum": sum, "min": min, "max": max, "round": round,
                "int": int, "float": float, "str": str, "bool": bool,
                "list": list, "dict": dict, "set": set, "tuple": tuple,
                "sorted": sorted, "enumerate": enumerate, "zip": zip,
                "range": range, "print": print, "abs": abs, "any": any, "all": all,
                "isinstance": isinstance, "hasattr": hasattr, "getattr": getattr,
            },
            "datetime": datetime,
            "pd": pd,
            "users": users,
            "teams": teams,
            "employees": employees,
            "projects": projects,
            "billing_entries": billing_entries,
            "staffing_entries": staffing_entries,
            "billing_rates": billing_rates,
            "cost_rates": cost_rates,
        }
        safe_locals = {}
        
        try:
            exec(generated_code, safe_globals, safe_locals)
            processed_data = safe_locals.get("result", [])
        except Exception as code_error:
            app.logger.error(f"Code execution error: {str(code_error)}")
            processed_data = {"error": f"Code execution failed: {str(code_error)}"}
        
        # ===== AGENT 3: Analysis =====
        progress_steps.append({"agent": "Agent 3", "status": "Analyzing results..."})
        agent3_prompt = f"""You are Agent 3: Analysis Agent.

User Question: "{user_query}"
Output Type: {query_plan.get('output_type', 'table')}
Processed Data:
{json.dumps(processed_data, indent=2, default=str)[:3000]}

IMPORTANT Context:
- Each billing_entry represents ONE DAY of billable work on a project (not a separate case/project)
- When you see "count" in the data, interpret it as "days worked" or "billing days"
- The same project/case_code can have multiple billing entries (one per day of work)

Based on the processed data and user's question:
- If output_type is "table": Format the data as a clear table with proper context
- If output_type is "insights": Analyze the data and provide key insights with correct terminology

Provide your response:"""

        response3 = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": agent3_prompt}],
            max_tokens=1500,
            temperature=0.5
        )
        
        analysis_result = response3.choices[0].message.content
        
        # ===== AGENT 4: Formatting =====
        progress_steps.append({"agent": "Agent 4", "status": "Formatting answer..."})
        agent4_prompt = f"""You are Agent 4: Formatting Agent.

Your job is to format the answer as clean, readable PLAIN TEXT with:
- Clear section titles (but NO markdown headers like ## or #)
- Well-formatted tables (use markdown table syntax | col1 | col2 |)
- Bullet points using - or • for insights
- Proper spacing and line breaks for readability
- Bold text using **text** if needed for emphasis

DO NOT USE: Markdown headers (##, ###, #) - they show up as literal text
USE INSTEAD: Plain text section titles with blank lines for separation

Original Question: "{user_query}"
Analysis Result:
{analysis_result}

Format this into a polished, professional response with proper plain text formatting:"""

        response4 = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": agent4_prompt}],
            max_tokens=2000,
            temperature=0.3
        )
        
        final_answer = response4.choices[0].message.content
        progress_steps.append({"agent": "Complete", "status": "Done"})
        
        return jsonify({
            "success": True,
            "message": final_answer,
            "progress": progress_steps
        })
        
    except Exception as e:
        app.logger.error(f"Chat API error: {str(e)}")
        return jsonify({"error": f"Error processing chat request: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
